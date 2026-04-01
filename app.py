import streamlit as st
import pandas as pd
import numpy as np
import re
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from scipy import stats
from itertools import combinations
from io import BytesIO
import numbers

try:
    import pingouin as pg
    HAS_PINGOUIN = True
except ImportError:
    HAS_PINGOUIN = False

# Publication-quality defaults — white background for papers/export
plt.rcParams.update({
    'font.family': 'Arial',
    'font.size': 10,
    'axes.linewidth': 1,
    'axes.labelcolor': 'black',
    'axes.titlecolor': 'black',
    'xtick.major.width': 1,
    'ytick.major.width': 1,
    'xtick.color': 'black',
    'ytick.color': 'black',
    'text.color': 'black',
    'legend.labelcolor': 'black',
    'figure.dpi': 300,
    'figure.facecolor': 'white',
    'axes.facecolor': 'white',
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'savefig.facecolor': 'white',
    'pdf.fonttype': 42,
    'ps.fonttype': 42,
    'svg.fonttype': 'none',
})

COLORS = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
STATES = ['NREM', 'REM', 'WAKE']
LIGHT_HOURS = list(range(12))
DARK_HOURS = list(range(12, 24))
MARKER_MAP = {'Circle': 'o', 'Square': 's', 'Triangle': '^', 'Diamond': 'D'}

# ─── Per-plot-type statistical test menus ────────────────────────────────────
# Bar chart: Factor1=Group, Factor2=Period(24h/Light/Dark)
BAR_TESTS_1F = ['None', 'Unpaired t-test', "Welch's t-test", 'Mann-Whitney U',
                'Paired t-test', 'Wilcoxon', 'One-way ANOVA', 'Kruskal-Wallis']
BAR_TESTS_2F = ['Two-way ANOVA', 'Two-way RM ANOVA', 'Mixed-effects model']
BAR_POSTHOC = ['Sidak', 'Tukey', 'Bonferroni', 'Dunnett']

# Time course: Factor1=Group, Factor2=Timepoint(0-23h)
TC_TESTS_1F = ['None', 'Unpaired t-test', 'Mann-Whitney U']
TC_TESTS_2F = ['Two-way ANOVA', 'Two-way RM ANOVA', 'Mixed-effects model']
TC_POSTHOC = ['Sidak', 'Tukey', 'Bonferroni']

# Delta/Spectrum: Factor1=Group, Factor2=Timepoint or Frequency
CURVE_TESTS_1F = ['None', 'Unpaired t-test', "Welch's t-test", 'Mann-Whitney U',
                  'Paired t-test', 'Wilcoxon']
CURVE_TESTS_2F = ['Two-way ANOVA', 'Two-way RM ANOVA', 'Mixed-effects model']
CURVE_POSTHOC = ['Sidak', 'Tukey', 'Bonferroni']

CORRECTIONS = ['None', 'Sidak', 'Bonferroni', 'FDR']


def _next_group_id():
    if '_group_id_counter' not in st.session_state:
        st.session_state._group_id_counter = 0
    st.session_state._group_id_counter += 1
    return st.session_state._group_id_counter


# ─── Mouse ID Range Parser ──────────────────────────────────────────────────

def parse_mouse_ids(text, valid_ids):
    if not text.strip():
        return [], []
    valid_set = set(valid_ids)
    tokens = re.split(r'[,\s]+', text.strip())
    matched, errors = [], []
    for token in tokens:
        token = token.strip()
        if not token:
            continue
        range_match = re.match(r'^([a-zA-Z]+)(\d+)\s*-\s*([a-zA-Z]*)(\d+)$', token)
        if range_match:
            prefix = range_match.group(1)
            start, end = int(range_match.group(2)), int(range_match.group(4))
            if start > end:
                start, end = end, start
            for n in range(start, end + 1):
                mid = f"{prefix}{n}"
                if mid in valid_set and mid not in matched:
                    matched.append(mid)
        elif token in valid_set:
            if token not in matched:
                matched.append(token)
        else:
            errors.append(token)
    return matched, errors


# ─── Data Loading ─────────────────────────────────────────────────────────────

@st.cache_data
def load_and_parse(uploaded_file):
    xls = pd.read_excel(uploaded_file, sheet_name=None)
    mouse_ids = [c for c in xls['Time'].columns if c not in ('idx', 'NR/R/W')]
    parsed = {}
    for sheet_name, df in xls.items():
        df = df.reset_index(drop=True)
        header_rows = df.index[
            df['NR/R/W'].isna() & df['idx'].astype(str).str.contains('BASAL', na=False)
        ].tolist()
        sheet_data = {}
        for hi, hrow in enumerate(header_rows):
            basal_name = str(df.iloc[hrow, 0])
            if hi + 1 < len(header_rows):
                end = header_rows[hi + 1]
            else:
                remaining = df.iloc[hrow + 1:]
                nan_rows = remaining.index[remaining['NR/R/W'].isna()]
                end = nan_rows[0] if len(nan_rows) > 0 else len(df)
            block = df.iloc[hrow + 1: end].copy()
            block_data = {}
            for state in STATES:
                state_rows = block[block['NR/R/W'] == state].copy()
                if len(state_rows) == 0:
                    continue
                idx_vals = state_rows['idx'].values
                if sheet_name.startswith('Spectrum'):
                    state_rows = state_rows.copy()
                    state_rows['idx'] = [int(str(v).replace('Hz', '')) for v in idx_vals]
                else:
                    state_rows['idx'] = [int(v) for v in idx_vals]
                state_rows = state_rows.set_index('idx').drop(columns=['NR/R/W'])
                state_rows = state_rows.apply(pd.to_numeric, errors='coerce')
                block_data[state] = state_rows
            sheet_data[basal_name] = block_data
        parsed[sheet_name] = sheet_data
    return parsed, mouse_ids


# ─── Aggregation ─────────────────────────────────────────────────────────────

def average_across_basals(sheet_data, state, mouse_ids, exclude_zero=False):
    frames = []
    for basal_name in sorted(sheet_data.keys()):
        if state not in sheet_data[basal_name]:
            continue
        df = sheet_data[basal_name][state][mouse_ids].copy()
        if exclude_zero:
            df = df.replace(0, np.nan)
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    stacked = np.stack([f.values for f in frames], axis=0)
    avg = np.nanmean(stacked, axis=0) if exclude_zero else np.mean(stacked, axis=0)
    return pd.DataFrame(avg, index=frames[0].index, columns=mouse_ids)


def group_mean_sem(df, error_type='SEM'):
    mean = df.mean(axis=1, skipna=True)
    err = df.std(axis=1, skipna=True) if error_type == 'SD' else df.sem(axis=1, skipna=True)
    return mean, err


def compute_total_time(time_sheet_data, mouse_ids, hours):
    result = {}
    for state in STATES:
        avg_df = average_across_basals(time_sheet_data, state, mouse_ids)
        if avg_df.empty:
            result[state] = np.zeros(len(mouse_ids))
            continue
        result[state] = avg_df.loc[avg_df.index.isin(hours)].sum(axis=0).values
    return result


# ─── Statistics Engine ───────────────────────────────────────────────────────

def p_to_stars(p, alpha=0.05):
    try:
        p = float(p)
        if np.isnan(p):
            return '-'
    except (TypeError, ValueError):
        return '-'
    if p <= 0.0001: return '****'
    if p <= 0.001: return '***'
    if p <= 0.01: return '**'
    if p <= alpha: return '*'
    return 'ns'


def format_p(p, display_mode, show_ns, alpha=0.05):
    stars = p_to_stars(p, alpha)
    if stars == 'ns' and not show_ns:
        return None
    if display_mode == 'Stars':
        return stars
    elif display_mode == 'Exact p':
        return f'p={p:.4f}' if p >= 0.0001 else 'p<0.0001'
    else:  # Both
        pstr = f'p={p:.4f}' if p >= 0.0001 else 'p<0.0001'
        return f'{stars}\n{pstr}'


def sidak_correction(p_raw, k):
    """Šídák correction: adjusted_p = 1 - (1 - p_raw)^k"""
    try:
        p = float(p_raw)
        if np.isnan(p) or k <= 0:
            return p
        return min(1.0, 1.0 - (1.0 - p) ** k)
    except (TypeError, ValueError):
        return np.nan


def apply_correction(p_vals, method):
    p_arr = np.array(p_vals, dtype=float)
    valid = ~np.isnan(p_arr)
    if not np.any(valid):
        return p_arr
    n = np.sum(valid)
    if method == 'Bonferroni':
        p_arr[valid] = np.minimum(p_arr[valid] * n, 1.0)
    elif method == 'Sidak':
        # Šídák: adjusted_p = 1 - (1 - p)^k, k = number of comparisons
        k = int(n)
        for i in range(len(p_arr)):
            if valid[i]:
                p_arr[i] = sidak_correction(p_arr[i], k)
    elif method == 'FDR':
        valid_p = p_arr[valid]
        order = np.argsort(valid_p)
        ranks = np.empty_like(order)
        ranks[order] = np.arange(1, len(valid_p) + 1)
        adjusted = np.minimum(valid_p * len(valid_p) / ranks, 1.0)
        order_desc = np.argsort(-ranks)
        for i in range(1, len(adjusted)):
            idx = order_desc[i]
            prev_idx = order_desc[i - 1]
            adjusted[idx] = min(adjusted[idx], adjusted[prev_idx])
        p_arr[valid] = adjusted
    return p_arr


def _clean_pair(a, b):
    """Clean two arrays: remove NaN, return as float."""
    a = np.array(a, dtype=float)
    b = np.array(b, dtype=float)
    a = a[~np.isnan(a)]
    b = b[~np.isnan(b)]
    return a, b


def _cohens_d(a, b):
    """Cohen's d effect size for two independent samples."""
    na, nb = len(a), len(b)
    if na < 2 or nb < 2:
        return np.nan
    pooled_std = np.sqrt(((na - 1) * np.var(a, ddof=1) + (nb - 1) * np.var(b, ddof=1)) / (na + nb - 2))
    return (np.mean(a) - np.mean(b)) / pooled_std if pooled_std > 0 else np.nan


def _safe_float(val):
    """Safely convert to float, returns np.nan on failure."""
    try:
        f = float(val)
        return f if not pd.isna(f) else np.nan
    except (TypeError, ValueError):
        return np.nan


def _get_col(row, *candidates, default=np.nan):
    """Get first available column value from a pingouin result row."""
    for col in candidates:
        if col in row.index:
            val = row[col]
            try:
                if not pd.isna(val):
                    return val
            except (TypeError, ValueError):
                return val  # non-numeric, return as-is
    return default


def _get_p_value(row):
    """Extract p-value from a pingouin result row, trying all known column names.

    Pingouin 0.5.x uses hyphens (p-unc), 0.6.x uses underscores (p_unc).
    We try both conventions.
    """
    # Try all known pingouin p-value column names (both hyphen and underscore variants)
    for col in ('p-unc', 'p_unc', 'p-GG-corr', 'p_GG_corr',
                'p-adjust', 'p_adjust', 'p-corr', 'p_corr',
                'p-val', 'p_val', 'p-tukey', 'p_tukey',
                'pval', 'p'):
        if col in row.index:
            val = row[col]
            try:
                if not pd.isna(val):
                    return float(val)
            except (TypeError, ValueError):
                continue
    # Fallback: search for any column containing 'p' with a value between 0 and 1
    for col in row.index:
        col_lower = col.lower()
        if ('p_' in col_lower or 'p-' in col_lower or col_lower == 'p') \
                and col not in ('np2', 'power', 'np2'):
            try:
                val = float(row[col])
                if 0 <= val <= 1 and not pd.isna(val):
                    return val
            except (TypeError, ValueError):
                continue
    return np.nan


def run_pairwise(data_lists, test_name, group_names, posthoc='Tukey'):
    """Run pairwise comparisons using pingouin. Returns list of result dicts."""
    results = []
    n = len(data_lists)
    if n < 2:
        return results

    # ── Build long-format for pingouin ──
    vals_all, grp_all = [], []
    for gi, d in enumerate(data_lists):
        arr = np.array(d, dtype=float)
        arr = arr[~np.isnan(arr)]
        vals_all.extend(arr.tolist())
        grp_all.extend([group_names[gi]] * len(arr))
    df_long = pd.DataFrame({'value': vals_all, 'group': grp_all})

    if len(df_long) < 4 or df_long['group'].nunique() < 2:
        return results

    # ── One-factor omnibus tests ──
    if test_name in ('One-way ANOVA', 'Kruskal-Wallis'):
        if HAS_PINGOUIN:
            try:
                if test_name == 'One-way ANOVA':
                    aov = pg.anova(data=df_long, dv='value', between='group')
                    row0 = aov.iloc[0]
                    results.append({
                        'Comparison': 'Omnibus F-test',
                        'Statistic': _safe_float(_get_col(row0, 'F')),
                        'df': f"{_get_col(row0, 'ddof1', 'DF', default='')}, {_get_col(row0, 'ddof2', 'DF_resid', default='')}",
                        'p': _get_p_value(row0),
                        'Effect size': _safe_float(_get_col(row0, 'np2', 'eta-sq')),
                        'Test': 'One-way ANOVA',
                    })
                else:
                    kw = pg.kruskal(data=df_long, dv='value', between='group')
                    row0 = kw.iloc[0]
                    results.append({
                        'Comparison': 'Omnibus H-test',
                        'Statistic': _safe_float(_get_col(row0, 'H')),
                        'df': str(_get_col(row0, 'ddof1', 'dof', default='')),
                        'p': _get_p_value(row0),
                        'Effect size': np.nan,
                        'Test': 'Kruskal-Wallis',
                    })
            except Exception as e:
                results.append({'Comparison': f'Error: {e}', 'Statistic': np.nan,
                                'df': '', 'p': np.nan, 'Effect size': np.nan,
                                'Test': test_name})
        else:
            # scipy fallback
            clean_all = [np.array(d, dtype=float) for d in data_lists]
            clean_all = [d[~np.isnan(d)] for d in clean_all if len(d[~np.isnan(d)]) > 0]
            if len(clean_all) >= 2:
                if test_name == 'One-way ANOVA':
                    f_stat, p = stats.f_oneway(*clean_all)
                    results.append({'Comparison': 'Omnibus F-test', 'Statistic': f_stat,
                                    'df': f'{len(clean_all)-1}, {sum(len(d) for d in clean_all)-len(clean_all)}',
                                    'p': p, 'Effect size': np.nan, 'Test': test_name})
                else:
                    h_stat, p = stats.kruskal(*clean_all)
                    results.append({'Comparison': 'Omnibus H-test', 'Statistic': h_stat,
                                    'df': str(len(clean_all) - 1),
                                    'p': p, 'Effect size': np.nan, 'Test': test_name})

        # Post-hoc pairwise
        if HAS_PINGOUIN and n >= 2:
            try:
                padjust_map = {'Tukey': 'none', 'Bonferroni': 'bonf',
                               'Sidak': 'sidak', 'Dunnett': 'none'}
                padjust_val = padjust_map.get(posthoc, 'none')

                if test_name == 'One-way ANOVA' and posthoc == 'Tukey':
                    ph = pg.pairwise_tukey(data=df_long, dv='value', between='group')
                    for _, row in ph.iterrows():
                        a_name, b_name = str(row['A']), str(row['B'])
                        a_arr, b_arr = _clean_pair(
                            data_lists[group_names.index(a_name)] if a_name in group_names else [],
                            data_lists[group_names.index(b_name)] if b_name in group_names else [])
                        results.append({
                            'Comparison': f"{a_name} vs {b_name}",
                            'Statistic': _safe_float(_get_col(row, 'T', 'q')),
                            'df': str(_get_col(row, 'dof', 'df', default='')),
                            'p': _get_p_value(row),
                            'Effect size': _cohens_d(a_arr, b_arr),
                            'Test': 'Tukey post-hoc',
                            'g1': a_name, 'g2': b_name,
                        })
                else:
                    ph = pg.pairwise_tests(data=df_long, dv='value', between='group',
                                           padjust=padjust_val)
                    for _, row in ph.iterrows():
                        a_name, b_name = str(row['A']), str(row['B'])
                        a_arr, b_arr = _clean_pair(
                            data_lists[group_names.index(a_name)] if a_name in group_names else [],
                            data_lists[group_names.index(b_name)] if b_name in group_names else [])
                        results.append({
                            'Comparison': f"{a_name} vs {b_name}",
                            'Statistic': _safe_float(_get_col(row, 'T', 'U-val', 'U_val')),
                            'df': str(_get_col(row, 'dof', 'df', default='')),
                            'p': _get_p_value(row),
                            'Effect size': _cohens_d(a_arr, b_arr),
                            'Test': f'{posthoc} post-hoc',
                            'g1': a_name, 'g2': b_name,
                        })
            except Exception:
                # Fallback: scipy pairwise
                for i, j in combinations(range(n), 2):
                    a, b = _clean_pair(data_lists[i], data_lists[j])
                    if len(a) >= 2 and len(b) >= 2:
                        t_stat, p = stats.ttest_ind(a, b)
                        results.append({
                            'Comparison': f"{group_names[i]} vs {group_names[j]}",
                            'Statistic': t_stat, 'df': str(len(a) + len(b) - 2),
                            'p': p, 'Effect size': _cohens_d(a, b),
                            'Test': f'{posthoc} post-hoc',
                            'g1': group_names[i], 'g2': group_names[j],
                        })
        return results

    # ── Simple two-sample tests (pingouin-based) ──
    for i, j in combinations(range(n), 2):
        a, b = _clean_pair(data_lists[i], data_lists[j])
        if len(a) < 2 or len(b) < 2:
            results.append({'Comparison': f"{group_names[i]} vs {group_names[j]}",
                            'Statistic': np.nan, 'df': '', 'p': np.nan,
                            'Effect size': np.nan,
                            'Test': test_name, 'g1': group_names[i], 'g2': group_names[j]})
            continue

        ef = _cohens_d(a, b)

        if HAS_PINGOUIN:
            try:
                if test_name == 'Unpaired t-test':
                    res = pg.ttest(a, b, paired=False, correction=False)
                    row0 = res.iloc[0]
                    results.append({
                        'Comparison': f"{group_names[i]} vs {group_names[j]}",
                        'Statistic': _safe_float(_get_col(row0, 'T')),
                        'df': str(_get_col(row0, 'dof', default='')),
                        'p': _get_p_value(row0),
                        'Effect size': _safe_float(_get_col(row0, 'cohen-d', 'cohen_d', default=ef)),
                        'Test': test_name, 'g1': group_names[i], 'g2': group_names[j],
                    })
                    continue
                elif test_name == "Welch's t-test":
                    res = pg.ttest(a, b, paired=False, correction=True)
                    row0 = res.iloc[0]
                    results.append({
                        'Comparison': f"{group_names[i]} vs {group_names[j]}",
                        'Statistic': _safe_float(_get_col(row0, 'T')),
                        'df': str(_get_col(row0, 'dof', default='')),
                        'p': _get_p_value(row0),
                        'Effect size': _safe_float(_get_col(row0, 'cohen-d', 'cohen_d', default=ef)),
                        'Test': test_name, 'g1': group_names[i], 'g2': group_names[j],
                    })
                    continue
                elif test_name == 'Mann-Whitney U':
                    res = pg.mwu(a, b)
                    row0 = res.iloc[0]
                    results.append({
                        'Comparison': f"{group_names[i]} vs {group_names[j]}",
                        'Statistic': _safe_float(_get_col(row0, 'U-val', 'U_val')),
                        'df': '',
                        'p': _get_p_value(row0),
                        'Effect size': _safe_float(_get_col(row0, 'RBC', 'CLES', default=ef)),
                        'Test': test_name, 'g1': group_names[i], 'g2': group_names[j],
                    })
                    continue
                elif test_name == 'Paired t-test':
                    min_n = min(len(a), len(b))
                    res = pg.ttest(a[:min_n], b[:min_n], paired=True)
                    row0 = res.iloc[0]
                    results.append({
                        'Comparison': f"{group_names[i]} vs {group_names[j]}",
                        'Statistic': _safe_float(_get_col(row0, 'T')),
                        'df': str(_get_col(row0, 'dof', default='')),
                        'p': _get_p_value(row0),
                        'Effect size': _safe_float(_get_col(row0, 'cohen-d', 'cohen_d', default=ef)),
                        'Test': test_name, 'g1': group_names[i], 'g2': group_names[j],
                    })
                    continue
                elif test_name == 'Wilcoxon':
                    min_n = min(len(a), len(b))
                    res = pg.wilcoxon(a[:min_n], b[:min_n])
                    row0 = res.iloc[0]
                    results.append({
                        'Comparison': f"{group_names[i]} vs {group_names[j]}",
                        'Statistic': _safe_float(_get_col(row0, 'W-val', 'W_val')),
                        'df': '',
                        'p': _get_p_value(row0),
                        'Effect size': _safe_float(_get_col(row0, 'RBC', 'CLES', default=ef)),
                        'Test': test_name, 'g1': group_names[i], 'g2': group_names[j],
                    })
                    continue
            except Exception:
                pass  # fall through to scipy

        # scipy fallback
        if test_name == 'Unpaired t-test':
            stat_val, p = stats.ttest_ind(a, b, equal_var=True)
            df_val = len(a) + len(b) - 2
        elif test_name == "Welch's t-test":
            stat_val, p = stats.ttest_ind(a, b, equal_var=False)
            s1, s2 = np.var(a, ddof=1), np.var(b, ddof=1)
            n1, n2 = len(a), len(b)
            num = (s1 / n1 + s2 / n2) ** 2
            den = (s1 / n1) ** 2 / (n1 - 1) + (s2 / n2) ** 2 / (n2 - 1)
            df_val = num / den if den > 0 else n1 + n2 - 2
        elif test_name == 'Mann-Whitney U':
            try:
                stat_val, p = stats.mannwhitneyu(a, b, alternative='two-sided')
            except ValueError:
                stat_val, p = np.nan, np.nan
            df_val = ''
        elif test_name == 'Paired t-test':
            min_n = min(len(a), len(b))
            stat_val, p = stats.ttest_rel(a[:min_n], b[:min_n])
            df_val = min_n - 1
        elif test_name == 'Wilcoxon':
            min_n = min(len(a), len(b))
            try:
                stat_val, p = stats.wilcoxon(a[:min_n], b[:min_n])
            except ValueError:
                stat_val, p = np.nan, np.nan
            df_val = ''
        else:
            stat_val, p = np.nan, np.nan
            df_val = ''

        results.append({
            'Comparison': f"{group_names[i]} vs {group_names[j]}",
            'Statistic': stat_val, 'df': str(df_val), 'p': p,
            'Effect size': ef, 'Test': test_name,
            'g1': group_names[i], 'g2': group_names[j],
        })

    return results


def _is_twoway(test_name):
    """Check if test is a two-factor test."""
    return test_name in ('Two-way ANOVA', 'Two-way RM ANOVA', 'Mixed-effects model')


def _manual_twoway_anova(long_df, dv, factor1, factor2):
    """Compute Two-way ANOVA (ordinary) manually using Type I SS.

    Returns dict with keys: 'table' (list of row dicts with Source/SS/DF/MS/F/p/np2),
                             'MSE' (residual mean square), 'DFe' (residual df).
    """
    from scipy.stats import f as f_dist

    data = long_df.dropna(subset=[dv]).copy()
    grand_mean = data[dv].mean()
    N = len(data)

    # Get factor levels
    levels_1 = data[factor1].unique()
    levels_2 = data[factor2].unique()
    a = len(levels_1)  # number of levels factor1
    b = len(levels_2)  # number of levels factor2

    # Cell means and counts
    cell_stats = data.groupby([factor1, factor2])[dv].agg(['mean', 'count', 'sum'])

    # SS Total
    SS_total = np.sum((data[dv].values - grand_mean) ** 2)

    # SS Factor1 (row factor, e.g. period/gene)
    SS_f1 = 0
    for lev in levels_1:
        group = data[data[factor1] == lev][dv]
        SS_f1 += len(group) * (group.mean() - grand_mean) ** 2

    # SS Factor2 (column factor, e.g. group/genotype)
    SS_f2 = 0
    for lev in levels_2:
        group = data[data[factor2] == lev][dv]
        SS_f2 += len(group) * (group.mean() - grand_mean) ** 2

    # SS Interaction
    SS_cells = 0
    for l1 in levels_1:
        for l2 in levels_2:
            cell = data[(data[factor1] == l1) & (data[factor2] == l2)][dv]
            if len(cell) > 0:
                SS_cells += len(cell) * (cell.mean() - grand_mean) ** 2
    SS_inter = SS_cells - SS_f1 - SS_f2

    # SS Residual (Error)
    SS_error = SS_total - SS_f1 - SS_f2 - SS_inter

    # Degrees of freedom
    df_f1 = a - 1
    df_f2 = b - 1
    df_inter = (a - 1) * (b - 1)
    df_error = N - a * b
    df_total = N - 1

    # Mean Squares
    MS_f1 = SS_f1 / df_f1 if df_f1 > 0 else 0
    MS_f2 = SS_f2 / df_f2 if df_f2 > 0 else 0
    MS_inter = SS_inter / df_inter if df_inter > 0 else 0
    MS_error = SS_error / df_error if df_error > 0 else 0

    # F statistics and p-values
    F_f1 = MS_f1 / MS_error if MS_error > 0 else np.nan
    F_f2 = MS_f2 / MS_error if MS_error > 0 else np.nan
    F_inter = MS_inter / MS_error if MS_error > 0 else np.nan

    p_f1 = 1 - f_dist.cdf(F_f1, df_f1, df_error) if not np.isnan(F_f1) else np.nan
    p_f2 = 1 - f_dist.cdf(F_f2, df_f2, df_error) if not np.isnan(F_f2) else np.nan
    p_inter = 1 - f_dist.cdf(F_inter, df_inter, df_error) if not np.isnan(F_inter) else np.nan

    # Partial eta-squared
    np2_f1 = SS_f1 / (SS_f1 + SS_error) if (SS_f1 + SS_error) > 0 else np.nan
    np2_f2 = SS_f2 / (SS_f2 + SS_error) if (SS_f2 + SS_error) > 0 else np.nan
    np2_inter = SS_inter / (SS_inter + SS_error) if (SS_inter + SS_error) > 0 else np.nan

    table = [
        {'Source': factor1, 'SS': SS_f1, 'DF': df_f1, 'MS': MS_f1,
         'F': F_f1, 'p': p_f1, 'np2': np2_f1},
        {'Source': factor2, 'SS': SS_f2, 'DF': df_f2, 'MS': MS_f2,
         'F': F_f2, 'p': p_f2, 'np2': np2_f2},
        {'Source': f'{factor1} x {factor2}', 'SS': SS_inter, 'DF': df_inter,
         'MS': MS_inter, 'F': F_inter, 'p': p_inter, 'np2': np2_inter},
        {'Source': 'Residual', 'SS': SS_error, 'DF': df_error, 'MS': MS_error,
         'F': np.nan, 'p': np.nan, 'np2': np.nan},
    ]

    return {'table': table, 'MSE': MS_error, 'DFe': df_error}


def _sidak_posthoc_with_pooled_mse(long_df, dv, row_factor, col_factor, MSE, DFe):
    """Šídák post-hoc: within each row level, compare column levels pairwise.

    Uses pooled MSE from the overall Two-way ANOVA.
    t = (mean1 - mean2) / sqrt(MSE * (1/n1 + 1/n2))
    df = DFe (residual degrees of freedom)
    Šídák correction: adjusted_p = 1 - (1 - p_raw)^k, k = C(n_cols, 2)
    """
    from scipy.stats import t as t_dist

    results = []
    col_levels = sorted(long_df[col_factor].unique())
    row_levels = sorted(long_df[row_factor].unique())
    n_cols = len(col_levels)
    k = n_cols * (n_cols - 1) // 2  # number of pairwise comparisons per family

    for row_lev in row_levels:
        row_data = long_df[long_df[row_factor] == row_lev]

        for i, j in combinations(range(n_cols), 2):
            g1_name = col_levels[i]
            g2_name = col_levels[j]
            g1_vals = row_data[row_data[col_factor] == g1_name][dv].dropna().values
            g2_vals = row_data[row_data[col_factor] == g2_name][dv].dropna().values

            n1, n2 = len(g1_vals), len(g2_vals)
            if n1 < 1 or n2 < 1 or MSE <= 0:
                results.append({
                    'Comparison': f'[{row_lev}] {g1_name} vs {g2_name}',
                    'Mean Diff': np.nan, 'Statistic': np.nan, 'df': str(DFe),
                    'p_raw': np.nan, 'p': np.nan,
                    'CI_low': np.nan, 'CI_high': np.nan,
                    'Significance': '-', 'Effect size': np.nan,
                    'Test': 'Sidak post-hoc',
                    'g1': str(g1_name), 'g2': str(g2_name),
                    'row_level': str(row_lev),
                })
                continue

            mean1, mean2 = np.mean(g1_vals), np.mean(g2_vals)
            mean_diff = mean1 - mean2
            se = np.sqrt(MSE * (1.0 / n1 + 1.0 / n2))
            t_stat = mean_diff / se if se > 0 else np.nan

            # Two-tailed p-value using residual df
            if not np.isnan(t_stat) and DFe > 0:
                p_raw = 2.0 * (1.0 - t_dist.cdf(abs(t_stat), DFe))
            else:
                p_raw = np.nan

            # Šídák correction: adjusted_p = 1 - (1 - p_raw)^k
            p_adj = sidak_correction(p_raw, k)

            # 95% CI for mean difference
            if DFe > 0 and se > 0:
                t_crit = t_dist.ppf(0.975, DFe)
                ci_low = mean_diff - t_crit * se
                ci_high = mean_diff + t_crit * se
            else:
                ci_low = ci_high = np.nan

            # Cohen's d using pooled MSE
            d = mean_diff / np.sqrt(MSE) if MSE > 0 else np.nan

            results.append({
                'Comparison': f'[{row_lev}] {g1_name} vs {g2_name}',
                'Mean Diff': mean_diff,
                'Statistic': t_stat,
                'df': str(DFe),
                'p_raw': p_raw,
                'p': p_adj,
                'CI_low': ci_low,
                'CI_high': ci_high,
                'Significance': p_to_stars(p_adj),
                'Effect size': d,
                'Test': 'Sidak post-hoc',
                'g1': str(g1_name), 'g2': str(g2_name),
                'row_level': str(row_lev),
            })

    return results


def run_twoway_anova(long_df, dv, between, within, test_type='Two-way ANOVA',
                     posthoc='Sidak', _export_store=None):
    """Run two-way ANOVA with Šídák post-hoc using pooled MSE.

    long_df must have columns: [dv, between, within, 'subject']
    between = column factor (e.g. 'group' / genotype)
    within  = row factor (e.g. 'period' / gene)

    Post-hoc: within each row level, compare column levels pairwise
    using pooled MSE from the overall ANOVA and Šídák correction.

    If _export_store is a dict, stores 'aov_result', 'posthoc_results',
    'long_df', 'dv', 'row_factor', 'col_factor' for Excel export.
    """
    results = []

    if long_df.empty or long_df[dv].dropna().shape[0] < 4:
        results.append({'Comparison': 'Error: insufficient data',
                        'Statistic': np.nan, 'df': '', 'p': np.nan,
                        'Effect size': np.nan, 'Test': test_type})
        return results

    try:
        if test_type == 'Two-way ANOVA':
            # ── Manual Two-way ANOVA (ordinary) ──
            aov_result = _manual_twoway_anova(long_df, dv,
                                              factor1=within,   # row factor
                                              factor2=between)  # column factor
            MSE = aov_result['MSE']
            DFe = aov_result['DFe']

            # Build ANOVA table results
            for row in aov_result['table']:
                f_val = row['F']
                p_val = row['p']
                results.append({
                    'Comparison': f"ANOVA: {row['Source']}",
                    'Statistic': f_val if not np.isnan(f_val) else np.nan,
                    'df': f"{row['DF']}, {DFe}" if row['Source'] != 'Residual' else str(row['DF']),
                    'SS': row['SS'],
                    'MS': row['MS'],
                    'p': p_val if not np.isnan(p_val) else np.nan,
                    'Effect size': row['np2'] if not np.isnan(row.get('np2', np.nan)) else np.nan,
                    'Test': test_type,
                })

            # ── Šídák post-hoc: within each row, compare columns ──
            posthoc_results = _sidak_posthoc_with_pooled_mse(
                long_df, dv, row_factor=within, col_factor=between,
                MSE=MSE, DFe=DFe)
            results.extend(posthoc_results)

            # Store for Excel export
            if isinstance(_export_store, dict):
                _export_store['aov_result'] = aov_result
                _export_store['posthoc_results'] = posthoc_results
                _export_store['long_df'] = long_df
                _export_store['dv'] = dv
                _export_store['row_factor'] = within
                _export_store['col_factor'] = between

        elif test_type == 'Two-way RM ANOVA' and HAS_PINGOUIN:
            aov = pg.rm_anova(data=long_df, dv=dv,
                              within=[between, within],
                              subject='subject')
            for _, row in aov.iterrows():
                src = _get_col(row, 'Source', default='')
                results.append({
                    'Comparison': f'ANOVA: {src}',
                    'Statistic': _safe_float(_get_col(row, 'F')),
                    'df': f"{_get_col(row, 'ddof1', 'DF', default='')}, {_get_col(row, 'ddof2', 'DF_resid', default='')}",
                    'p': _get_p_value(row),
                    'Effect size': _safe_float(_get_col(row, 'np2')),
                    'Test': test_type,
                })
            # RM post-hoc via pingouin
            try:
                ph = pg.pairwise_tests(data=long_df, dv=dv,
                                       within=[between, within],
                                       subject='subject', padjust='sidak')
                for _, row in ph.iterrows():
                    contrast = str(_get_col(row, 'Contrast', default=''))
                    a_val, b_val = str(_get_col(row, 'A', default='')), str(_get_col(row, 'B', default=''))
                    comp_label = f"[{contrast}] {a_val} vs {b_val}" if contrast and contrast != between else f"{a_val} vs {b_val}"
                    results.append({
                        'Comparison': comp_label,
                        'Statistic': _safe_float(_get_col(row, 'T')),
                        'df': str(_get_col(row, 'dof', 'df', default='')),
                        'p': _get_p_value(row),
                        'Effect size': _safe_float(_get_col(row, 'cohen', 'hedges', 'cohen-d', 'cohen_d')),
                        'Test': 'Sidak post-hoc', 'g1': a_val, 'g2': b_val,
                    })
            except Exception:
                pass

        elif test_type == 'Mixed-effects model' and HAS_PINGOUIN:
            aov = pg.mixed_anova(data=long_df, dv=dv,
                                 between=between, within=within,
                                 subject='subject')
            for _, row in aov.iterrows():
                src = _get_col(row, 'Source', default='')
                results.append({
                    'Comparison': f'ANOVA: {src}',
                    'Statistic': _safe_float(_get_col(row, 'F')),
                    'df': f"{_get_col(row, 'ddof1', 'DF1', default='')}, {_get_col(row, 'ddof2', 'DF2', default='')}",
                    'p': _get_p_value(row),
                    'Effect size': _safe_float(_get_col(row, 'np2')),
                    'Test': test_type,
                })
            try:
                ph = pg.pairwise_tests(data=long_df, dv=dv,
                                       between=between, within=within,
                                       subject='subject', padjust='sidak')
                for _, row in ph.iterrows():
                    contrast = str(_get_col(row, 'Contrast', default=''))
                    a_val, b_val = str(_get_col(row, 'A', default='')), str(_get_col(row, 'B', default=''))
                    comp_label = f"[{contrast}] {a_val} vs {b_val}" if contrast and contrast != between else f"{a_val} vs {b_val}"
                    results.append({
                        'Comparison': comp_label,
                        'Statistic': _safe_float(_get_col(row, 'T')),
                        'df': str(_get_col(row, 'dof', 'df', default='')),
                        'p': _get_p_value(row),
                        'Effect size': _safe_float(_get_col(row, 'cohen', 'hedges')),
                        'Test': 'Sidak post-hoc', 'g1': a_val, 'g2': b_val,
                    })
            except Exception:
                pass

        else:
            # Fallback: if pingouin not available for RM/Mixed, use manual two-way
            aov_result = _manual_twoway_anova(long_df, dv, factor1=within, factor2=between)
            MSE, DFe = aov_result['MSE'], aov_result['DFe']
            for row in aov_result['table']:
                results.append({
                    'Comparison': f"ANOVA: {row['Source']}",
                    'Statistic': row['F'], 'df': f"{row['DF']}, {DFe}" if row['Source'] != 'Residual' else str(row['DF']),
                    'SS': row['SS'], 'MS': row['MS'],
                    'p': row['p'], 'Effect size': row.get('np2', np.nan),
                    'Test': test_type,
                })
            posthoc_results = _sidak_posthoc_with_pooled_mse(
                long_df, dv, row_factor=within, col_factor=between, MSE=MSE, DFe=DFe)
            results.extend(posthoc_results)

    except Exception as e:
        results.append({'Comparison': f'Error: {e}', 'Statistic': np.nan,
                        'df': '', 'p': np.nan, 'Effect size': np.nan,
                        'Test': test_type})
    return results


def run_pointwise_test(data_lists, test_name, group_names):
    """Run a single omnibus or pairwise test at one point. Returns p-value."""
    clean = [np.array(d, dtype=float) for d in data_lists]
    clean = [d[~np.isnan(d)] for d in clean]
    clean = [d for d in clean if len(d) > 1]
    if len(clean) < 2:
        return np.nan

    if HAS_PINGOUIN and len(clean) == 2:
        try:
            if test_name == 'Unpaired t-test':
                return _get_p_value(pg.ttest(clean[0], clean[1], paired=False, correction=False).iloc[0])
            elif test_name == "Welch's t-test":
                return _get_p_value(pg.ttest(clean[0], clean[1], paired=False, correction=True).iloc[0])
            elif test_name == 'Mann-Whitney U':
                return _get_p_value(pg.mwu(clean[0], clean[1]).iloc[0])
            elif test_name == 'Paired t-test':
                mn = min(len(clean[0]), len(clean[1]))
                return _get_p_value(pg.ttest(clean[0][:mn], clean[1][:mn], paired=True).iloc[0])
            elif test_name == 'Wilcoxon':
                mn = min(len(clean[0]), len(clean[1]))
                return _get_p_value(pg.wilcoxon(clean[0][:mn], clean[1][:mn]).iloc[0])
        except Exception:
            pass  # fall through to scipy

    if test_name in ('Unpaired t-test', "Welch's t-test") and len(clean) == 2:
        eq_var = test_name == 'Unpaired t-test'
        _, p = stats.ttest_ind(clean[0], clean[1], equal_var=eq_var)
        return p
    elif test_name == 'Mann-Whitney U' and len(clean) == 2:
        try:
            _, p = stats.mannwhitneyu(clean[0], clean[1], alternative='two-sided')
        except ValueError:
            return np.nan
        return p
    elif test_name == 'Paired t-test' and len(clean) == 2:
        min_n = min(len(clean[0]), len(clean[1]))
        _, p = stats.ttest_rel(clean[0][:min_n], clean[1][:min_n])
        return p
    elif test_name == 'Wilcoxon' and len(clean) == 2:
        min_n = min(len(clean[0]), len(clean[1]))
        try:
            _, p = stats.wilcoxon(clean[0][:min_n], clean[1][:min_n])
        except ValueError:
            return np.nan
        return p
    elif test_name in ('One-way ANOVA', 'Kruskal-Wallis'):
        if test_name == 'One-way ANOVA':
            _, p = stats.f_oneway(*clean)
        else:
            _, p = stats.kruskal(*clean)
        return p
    elif test_name.startswith('Two-way') or test_name == 'Mixed-effects model':
        # For pointwise, fall back to one-way test
        if len(clean) >= 2:
            _, p = stats.f_oneway(*clean)
            return p
    return np.nan


def add_significance_bracket(ax, x1, x2, y, text, lw=1.0, fs=8):
    h = (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.02
    ax.plot([x1, x1, x2, x2], [y, y + h, y + h, y], color='black', lw=lw, clip_on=False)
    ax.text((x1 + x2) / 2, y + h, text, ha='center', va='bottom', fontsize=fs)


def add_significance_markers(ax, x_vals, p_vals, y_vals, stat_s):
    fs = stat_s.get('sig_font_size', 7)
    display = stat_s.get('sig_display', 'Stars')
    show_ns = stat_s.get('show_ns', False)
    alpha = stat_s.get('sig_alpha', 0.05)
    for x, p, y in zip(x_vals, p_vals, y_vals):
        if np.isnan(p):
            continue
        label = format_p(p, display, show_ns, alpha)
        if label is not None:
            ax.text(x, y, label, ha='center', va='bottom', fontsize=fs, color='black')


# ─── Plot Settings ───────────────────────────────────────────────────────────

def _init_plot_settings(key, groups, plot_type):
    sk = f'_ps_{key}'
    if sk not in st.session_state:
        st.session_state[sk] = {}
    s = st.session_state[sk]

    s.setdefault('y_auto', True)
    s.setdefault('y_min', 0.0)
    s.setdefault('y_max', 100.0)
    s.setdefault('y_tick', 0.0)
    s.setdefault('y_label', '')
    s.setdefault('y_tight', False)
    s.setdefault('x_auto', True)
    s.setdefault('x_min', 0.0)
    s.setdefault('x_max', 24.0)
    s.setdefault('x_tick', 0.0)
    s.setdefault('x_label', '')
    s.setdefault('x_tight', False)
    s.setdefault('bg_mode', 'Light/Dark')

    default_w = {'bar': 12.0, 'line': 14.0, 'spectrum': 14.0}.get(plot_type, 10.0)
    default_h = {'bar': 4.0, 'line': 4.0, 'spectrum': 4.0}.get(plot_type, 4.0)
    s.setdefault('fig_width', default_w)
    s.setdefault('fig_height', default_h)
    s.setdefault('hspace', 0.3)
    s.setdefault('wspace', 0.3)

    s.setdefault('point_size', 12.0)
    s.setdefault('point_fill', 'Solid')
    s.setdefault('point_shape', 'Circle')
    s.setdefault('line_width', 1.5)
    s.setdefault('line_style', 'Solid')

    s.setdefault('eb_style', 'Shade')
    s.setdefault('shade_alpha', 0.2)
    if 'shade_colors' not in s:
        s['shade_colors'] = {}
    s.setdefault('error_type', 'SEM')
    s.setdefault('error_capsize', 3.0)
    s.setdefault('error_linewidth', 1.0)
    s.setdefault('error_cap_mode', 'With cap')

    s.setdefault('bar_fill', 'Solid')
    s.setdefault('bar_edge_width', 1.5)

    s.setdefault('sig_height', 0.05)
    s.setdefault('sig_line_width', 1.0)
    s.setdefault('sig_font_size', 10)

    # Statistics (per-plot)
    s.setdefault('stat_test', 'None')
    s.setdefault('stat_posthoc', 'Sidak')
    s.setdefault('stat_correction', 'None')
    s.setdefault('sig_display', 'Stars')
    s.setdefault('show_ns', False)
    s.setdefault('sig_alpha', 0.05)

    s.setdefault('font_size', 10)

    if 'group_colors' not in s:
        s['group_colors'] = {}
    for gi, group in enumerate(groups):
        gid = group['gid']
        if gid not in s['group_colors']:
            s['group_colors'][gid] = COLORS[gi % len(COLORS)]
        if gid not in s['shade_colors']:
            s['shade_colors'][gid] = s['group_colors'][gid]

    return s


def _render_stat_tab(sk, s, plot_type, n_groups):
    """Render the Statistics tab content based on plot type."""
    # Select available tests per plot type
    if plot_type == 'bar':
        tests_1f = BAR_TESTS_1F
        tests_2f = BAR_TESTS_2F
        posthoc_opts = BAR_POSTHOC
        factor2_label = "Period (24h/Light/Dark)"
    elif plot_type == 'line':  # timecourse
        tests_1f = TC_TESTS_1F
        tests_2f = TC_TESTS_2F
        posthoc_opts = TC_POSTHOC
        factor2_label = "Time point"
    else:  # spectrum, delta
        tests_1f = CURVE_TESTS_1F
        tests_2f = CURVE_TESTS_2F
        posthoc_opts = CURVE_POSTHOC
        factor2_label = "Frequency / Time point"

    all_tests = tests_1f + tests_2f
    cur_test = s['stat_test'] if s['stat_test'] in all_tests else 'None'

    st.caption("ONE-FACTOR TESTS")
    s['stat_test'] = st.selectbox("Test", tests_1f, index=tests_1f.index(cur_test) if cur_test in tests_1f else 0,
                                  key=f"{sk}__st")

    if s['stat_test'] == 'None' and HAS_PINGOUIN:
        st.caption("TWO-FACTOR TESTS")
        st.caption(f"Factor 1 = Group, Factor 2 = {factor2_label}")
        twoway_opts = ['None'] + tests_2f
        cur_2f = s.get('stat_test_2f', 'None')
        if cur_2f not in twoway_opts:
            cur_2f = 'None'
        chosen_2f = st.selectbox("Two-factor test", twoway_opts,
                                 index=twoway_opts.index(cur_2f),
                                 key=f"{sk}__st2f")
        s['stat_test_2f'] = chosen_2f
        if chosen_2f != 'None':
            s['stat_test'] = chosen_2f
    elif s['stat_test'] == 'None':
        s['stat_test_2f'] = 'None'
    else:
        s['stat_test_2f'] = 'None'

    if s['stat_test'] in ('One-way ANOVA', 'Kruskal-Wallis') or s['stat_test'] in tests_2f:
        st.caption("POST-HOC")
        cur_ph = s['stat_posthoc'] if s['stat_posthoc'] in posthoc_opts else posthoc_opts[0]
        s['stat_posthoc'] = st.selectbox("Post-hoc", posthoc_opts,
                                         index=posthoc_opts.index(cur_ph),
                                         key=f"{sk}__sph")

    if s['stat_test'] != 'None':
        st.caption("OPTIONS")
        c1, c2, c3 = st.columns(3)
        with c1:
            s['stat_correction'] = st.selectbox("Correction", CORRECTIONS,
                                                index=CORRECTIONS.index(s['stat_correction']),
                                                key=f"{sk}__sc")
        with c2:
            displays = ['Stars', 'Exact p', 'Both']
            s['sig_display'] = st.radio("Display", displays,
                                        index=displays.index(s['sig_display']),
                                        key=f"{sk}__sd", horizontal=True)
        with c3:
            s['sig_alpha'] = st.number_input("α threshold", 0.001, 0.10, s['sig_alpha'], 0.005,
                                             key=f"{sk}__sa2", format="%.3f")
        s['show_ns'] = st.checkbox("Show ns", value=s['show_ns'], key=f"{sk}__sns")

        st.caption("ANNOTATION")
        s['sig_font_size'] = st.number_input("Star font size", 6, 20,
                                              s.get('sig_font_size', 10), 1,
                                              key=f"{sk}__sfs")

    if not HAS_PINGOUIN:
        st.warning("Install `pingouin` for Two-way ANOVA, RM ANOVA, Mixed-effects, and Tukey post-hoc. "
                   "Run: `pip install pingouin`")


def get_plot_settings(key, groups, plot_type='line', n_groups=None):
    s = _init_plot_settings(key, groups, plot_type)
    sk = f'_ps_{key}'
    if n_groups is None:
        n_groups = len(groups)

    with st.expander("Plot Settings", expanded=False):
        tabs = st.tabs(["Appearance", "Error Bars", "Axes & Layout", "Annotation", "Statistics"])

        # ═══ TAB 1: Appearance ═══
        with tabs[0]:
            st.caption("GROUP COLORS")
            color_cols = st.columns(max(len(groups), 1))
            for gi, group in enumerate(groups):
                gid = group['gid']
                with color_cols[gi]:
                    s['group_colors'][gid] = st.color_picker(
                        group['name'],
                        value=s['group_colors'].get(gid, COLORS[gi % len(COLORS)]),
                        key=f"{sk}__clr_{gid}",
                    )

            if plot_type == 'bar':
                st.caption("BAR STYLE")
                b1, b2 = st.columns(2)
                with b1:
                    s['bar_fill'] = st.radio("Fill", ['Solid', 'Hollow'],
                                             index=['Solid', 'Hollow'].index(s['bar_fill']),
                                             key=f"{sk}__bf", horizontal=True)
                with b2:
                    s['bar_edge_width'] = st.number_input("Border width", 0.5, 5.0,
                                                          s['bar_edge_width'], 0.5,
                                                          key=f"{sk}__be", format="%.1f")

            if plot_type in ('line', 'spectrum'):
                st.caption("LINE STYLE")
                l1, l2 = st.columns(2)
                with l1:
                    s['line_width'] = st.number_input("Width", 0.5, 5.0, s['line_width'], 0.25,
                                                      key=f"{sk}__lw", format="%.2f")
                with l2:
                    s['line_style'] = st.radio("Style", ['Solid', 'Dashed'],
                                               index=['Solid', 'Dashed'].index(s['line_style']),
                                               key=f"{sk}__ls", horizontal=True)

            st.caption("DATA POINTS")
            p1, p2, p3 = st.columns(3)
            with p1:
                s['point_size'] = st.number_input("Size", 1.0, 80.0, s['point_size'], 1.0,
                                                  key=f"{sk}__ps", format="%.0f")
            with p2:
                shapes = list(MARKER_MAP.keys())
                s['point_shape'] = st.selectbox("Shape", shapes,
                                                index=shapes.index(s['point_shape']),
                                                key=f"{sk}__pm")
            with p3:
                s['point_fill'] = st.radio("Fill", ['Solid', 'Hollow'],
                                           index=['Solid', 'Hollow'].index(s['point_fill']),
                                           key=f"{sk}__pf", horizontal=True)

            st.caption("TYPOGRAPHY")
            s['font_size'] = st.number_input("Font size", 4, 24, s['font_size'], 1, key=f"{sk}__fs")

        # ═══ TAB 2: Error Bars ═══
        with tabs[1]:
            st.caption("ERROR TYPE")
            s['error_type'] = st.radio("Type", ['SEM', 'SD'],
                                       index=['SEM', 'SD'].index(s['error_type']),
                                       key=f"{sk}__et", horizontal=True)

            st.caption("STYLE")
            if plot_type in ('line', 'spectrum'):
                eb_options = ['Line + cap', 'Line no cap', 'Shade']
                if s['eb_style'] == 'Shade':
                    cur_eb_idx = 2
                elif s.get('error_cap_mode', 'With cap') == 'No cap':
                    cur_eb_idx = 1
                else:
                    cur_eb_idx = 0
                eb_choice = st.radio("Error bar style", eb_options, index=cur_eb_idx,
                                     key=f"{sk}__ebs", horizontal=True)
                if eb_choice == 'Shade':
                    s['eb_style'] = 'Shade'
                    s['error_cap_mode'] = 'With cap'
                elif eb_choice == 'Line no cap':
                    s['eb_style'] = 'Line'
                    s['error_cap_mode'] = 'No cap'
                else:
                    s['eb_style'] = 'Line'
                    s['error_cap_mode'] = 'With cap'
            else:
                cap_options = ['With cap', 'No cap']
                s['error_cap_mode'] = st.radio("Cap style", cap_options,
                                               index=cap_options.index(s.get('error_cap_mode', 'With cap')),
                                               key=f"{sk}__ecm", horizontal=True)

            st.caption("PARAMETERS")
            e1, e2, e3 = st.columns(3)
            with e1:
                s['error_capsize'] = st.number_input("Cap size", 0.0, 10.0, s['error_capsize'], 0.5,
                                                     key=f"{sk}__ec", format="%.1f")
            with e2:
                s['error_linewidth'] = st.number_input("Line width", 0.5, 5.0, s['error_linewidth'], 0.25,
                                                       key=f"{sk}__ew", format="%.2f")
            with e3:
                s['shade_alpha'] = st.number_input("Shade alpha", 0.0, 1.0, s['shade_alpha'], 0.05,
                                                   key=f"{sk}__sa", format="%.2f")

            if plot_type in ('line', 'spectrum') and s['eb_style'] == 'Shade':
                st.caption("SHADE COLORS")
                shade_cols = st.columns(max(len(groups), 1))
                for gi, group in enumerate(groups):
                    gid = group['gid']
                    with shade_cols[gi]:
                        s['shade_colors'][gid] = st.color_picker(
                            f"{group['name']}",
                            value=s['shade_colors'].get(gid, s['group_colors'].get(gid, COLORS[gi % len(COLORS)])),
                            key=f"{sk}__shd_{gid}",
                        )

        # ═══ TAB 3: Axes & Layout ═══
        with tabs[2]:
            st.caption("Y AXIS")
            y1, y2, y3, y4 = st.columns(4)
            with y1:
                s['y_auto'] = st.checkbox("Auto", value=s['y_auto'], key=f"{sk}__ya")
            with y2:
                s['y_min'] = st.number_input("Min", value=s['y_min'], key=f"{sk}__ymn",
                                             format="%.2f", disabled=s['y_auto'])
            with y3:
                s['y_max'] = st.number_input("Max", value=s['y_max'], key=f"{sk}__ymx",
                                             format="%.2f", disabled=s['y_auto'])
            with y4:
                s['y_tick'] = st.number_input("Tick (0=auto)", value=s['y_tick'],
                                              min_value=0.0, key=f"{sk}__yt", format="%.2f")

            st.caption("X AXIS")
            x1, x2, x3, x4 = st.columns(4)
            with x1:
                s['x_auto'] = st.checkbox("Auto", value=s['x_auto'], key=f"{sk}__xa")
            with x2:
                s['x_min'] = st.number_input("Min", value=s['x_min'], key=f"{sk}__xmn",
                                             format="%.2f", disabled=s['x_auto'])
            with x3:
                s['x_max'] = st.number_input("Max", value=s['x_max'], key=f"{sk}__xmx",
                                             format="%.2f", disabled=s['x_auto'])
            with x4:
                s['x_tick'] = st.number_input("Tick (0=auto)", value=s['x_tick'],
                                              min_value=0.0, key=f"{sk}__xt", format="%.2f")

            st.caption("AXIS LABELS")
            lb1, lb2 = st.columns(2)
            with lb1:
                s['y_label'] = st.text_input("Y label", value=s['y_label'], key=f"{sk}__yl",
                                             placeholder="default")
            with lb2:
                s['x_label'] = st.text_input("X label", value=s['x_label'], key=f"{sk}__xl",
                                             placeholder="default")

            st.caption("OPTIONS")
            o1, o2, o3 = st.columns(3)
            with o1:
                s['y_tight'] = st.checkbox("Y tight (0→origin)", value=s['y_tight'], key=f"{sk}__ytight")
            with o2:
                s['x_tight'] = st.checkbox("X tight (0→origin)", value=s['x_tight'], key=f"{sk}__xtight")
            with o3:
                if plot_type in ('line', 'spectrum'):
                    s['bg_mode'] = st.radio("Background", ['Light/Dark', 'White'],
                                            index=['Light/Dark', 'White'].index(s['bg_mode']),
                                            key=f"{sk}__bg", horizontal=True)

            st.caption("FIGURE SIZE")
            f1, f2, f3, f4 = st.columns(4)
            with f1:
                s['fig_width'] = st.number_input("Width (in)", 3.0, 30.0, s['fig_width'], 0.5,
                                                 key=f"{sk}__fw", format="%.1f")
            with f2:
                s['fig_height'] = st.number_input("Height (in)", 1.5, 15.0, s['fig_height'], 0.5,
                                                  key=f"{sk}__fh", format="%.1f")
            with f3:
                s['hspace'] = st.number_input("V space", 0.0, 2.0, s['hspace'], 0.05,
                                              key=f"{sk}__hs", format="%.2f")
            with f4:
                s['wspace'] = st.number_input("H space", 0.0, 2.0, s['wspace'], 0.05,
                                              key=f"{sk}__ws", format="%.2f")

        # ═══ TAB 4: Annotation style ═══
        with tabs[3]:
            st.caption("ANNOTATION STYLE")
            s1, s2, s3 = st.columns(3)
            with s1:
                s['sig_height'] = st.number_input("Bracket offset", 0.01, 0.50, s['sig_height'], 0.01,
                                                  key=f"{sk}__sh", format="%.2f")
            with s2:
                s['sig_line_width'] = st.number_input("Line width", 0.5, 5.0, s['sig_line_width'], 0.25,
                                                      key=f"{sk}__sl", format="%.2f")
            with s3:
                s['sig_font_size'] = st.number_input("Star size", 4, 20, s['sig_font_size'], 1,
                                                     key=f"{sk}__sf")

        # ═══ TAB 5: Statistics ═══
        with tabs[4]:
            _render_stat_tab(sk, s, plot_type, n_groups)

    # Build output dict
    out = dict(s)
    out['colors'] = [s['group_colors'].get(g['gid'], COLORS[i % len(COLORS)]) for i, g in enumerate(groups)]
    out['shade_color_list'] = [s['shade_colors'].get(g['gid'], out['colors'][i]) for i, g in enumerate(groups)]
    out['y_min'] = None if s['y_auto'] else s['y_min']
    out['y_max'] = None if s['y_auto'] else s['y_max']
    out['x_min'] = None if s['x_auto'] else s['x_min']
    out['x_max'] = None if s['x_auto'] else s['x_max']
    out['marker'] = MARKER_MAP.get(s['point_shape'], 'o')
    out['linestyle'] = '-' if s.get('line_style', 'Solid') == 'Solid' else '--'
    out['effective_capsize'] = 0.0 if s.get('error_cap_mode', 'With cap') == 'No cap' else s['error_capsize']
    out['active_stat'] = s['stat_test'] if s['stat_test'] != 'None' else None
    return out


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _fix_yticks(ax, s):
    ax.figure.canvas.draw()
    ticks = ax.get_yticks()
    if len(ticks) < 2:
        return
    ymin, ymax = ax.get_ylim()
    visible_ticks = ticks[(ticks >= ymin) & (ticks <= ymax + (ticks[1] - ticks[0]) * 0.5)]
    if len(visible_ticks) > 0:
        new_ymax = visible_ticks[-1]
        if new_ymax > ymax:
            ax.set_ylim(ymin, new_ymax)
    if s.get('y_min') is not None and s.get('y_max') is not None:
        ax.set_ylim(s['y_min'], s['y_max'])


def apply_axis_settings(ax, s, default_xlabel='', default_ylabel=''):
    fs = s.get('font_size', 10)
    if s.get('y_min') is not None and s.get('y_max') is not None:
        ax.set_ylim(s['y_min'], s['y_max'])
    if s.get('x_min') is not None and s.get('x_max') is not None:
        ax.set_xlim(s['x_min'], s['x_max'])
    if s.get('y_tight', False):
        yl = ax.get_ylim()
        ax.set_ylim(bottom=0, top=yl[1])
    if s.get('x_tight', False):
        xl = ax.get_xlim()
        ax.set_xlim(left=0, right=xl[1])
    if s.get('y_tick', 0) > 0:
        ax.yaxis.set_major_locator(mticker.MultipleLocator(s['y_tick']))
    if s.get('x_tick', 0) > 0:
        ax.xaxis.set_major_locator(mticker.MultipleLocator(s['x_tick']))
    ylabel = s.get('y_label', '') or default_ylabel
    xlabel = s.get('x_label', '') or default_xlabel
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=fs)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=fs)
    ax.tick_params(labelsize=fs)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    _fix_yticks(ax, s)


def scatter_points(ax, x_pos, vals, color, s):
    marker = s.get('marker', 'o')
    size = s.get('point_size', 12)
    is_hollow = s.get('point_fill', 'Solid') == 'Hollow'
    n = len(vals)
    jitter = np.random.uniform(-0.08, 0.08, n)
    if is_hollow:
        ax.scatter(np.full(n, x_pos) + jitter, vals, facecolors='none',
                   edgecolors=color, s=size, marker=marker, zorder=5, linewidth=0.8)
    else:
        ax.scatter(np.full(n, x_pos) + jitter, vals, color=color,
                   edgecolors='white', s=size, marker=marker, zorder=5, linewidth=0.5)


def draw_error(ax, x, mean, sem, color, shade_color, s, gi):
    capsize = s.get('effective_capsize', s.get('error_capsize', 3))
    if s.get('eb_style', 'Shade') == 'Shade':
        ax.fill_between(x, mean - sem, mean + sem,
                        color=shade_color, alpha=s.get('shade_alpha', 0.2))
    else:
        ax.errorbar(x, mean, yerr=sem, fmt='none', ecolor=color,
                    capsize=capsize,
                    elinewidth=s.get('error_linewidth', 1.0),
                    capthick=s.get('error_linewidth', 1.0))


def add_light_dark_shading(ax, s):
    if s.get('bg_mode', 'Light/Dark') == 'White':
        return
    ymin, ymax = ax.get_ylim()
    ax.axvspan(-0.5, 11.5, alpha=0.1, color='gold', zorder=0)
    ax.axvspan(11.5, 23.5, alpha=0.1, color='gray', zorder=0)
    ax.set_ylim(ymin, ymax)


# ─── Export ──────────────────────────────────────────────────────────────────

def export_figure(fig, fmt, dpi=300):
    buf = BytesIO()
    if fmt == 'pdf':
        fig.savefig(buf, format='pdf', bbox_inches='tight')
    elif fmt == 'svg':
        fig.savefig(buf, format='svg', bbox_inches='tight')
    elif fmt == 'tiff':
        fig.savefig(buf, format='tiff', dpi=max(dpi, 600), bbox_inches='tight',
                    pil_kwargs={'compression': 'tiff_lzw'})
    else:
        fig.savefig(buf, format=fmt, dpi=dpi, bbox_inches='tight')
    buf.seek(0)
    return buf


def show_export_buttons(fig, key_prefix):
    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button("PDF", export_figure(fig, 'pdf'),
                           f"{key_prefix}.pdf", mime="application/pdf",
                           key=f"{key_prefix}_pdf")
    with col2:
        st.download_button("TIFF 600dpi", export_figure(fig, 'tiff'),
                           f"{key_prefix}.tiff", mime="image/tiff",
                           key=f"{key_prefix}_tiff")
    with col3:
        st.download_button("SVG", export_figure(fig, 'svg'),
                           f"{key_prefix}.svg", mime="image/svg+xml",
                           key=f"{key_prefix}_svg")


# ─── Plot: Bar Chart ────────────────────────────────────────────────────────

def plot_bar_chart(parsed, groups, s):
    time_data = parsed['Time']
    periods = {'24h': list(range(24)), 'Light': LIGHT_HOURS, 'Dark': DARK_HOURS}
    fs = s['font_size']
    capsize = s.get('effective_capsize', s.get('error_capsize', 3))
    stat_test = s.get('active_stat')
    posthoc = s.get('stat_posthoc', 'Sidak')
    stat_results_all = []

    fig, axes = plt.subplots(1, 3, figsize=(s['fig_width'], s['fig_height']))
    fig.subplots_adjust(wspace=s['wspace'])

    for si, state in enumerate(STATES):
        ax = axes[si]
        n_groups = len(groups)
        bar_width = 0.8 / n_groups
        period_names = list(periods.keys())
        n_periods = len(period_names)
        all_group_data = {pi: [] for pi in range(n_periods)}

        for gi, group in enumerate(groups):
            means, errs, all_vals = [], [], []
            for pname in period_names:
                vals = compute_total_time(time_data, group['ids'], periods[pname])[state]
                means.append(np.mean(vals))
                if s['error_type'] == 'SD':
                    errs.append(np.std(vals, ddof=1) if len(vals) > 1 else 0)
                else:
                    errs.append(np.std(vals, ddof=1) / np.sqrt(len(vals)) if len(vals) > 1 else 0)
                all_vals.append(vals)
            for pi_idx, v in enumerate(all_vals):
                all_group_data[pi_idx].append(v)

            x = np.arange(n_periods)
            offset = (gi - (n_groups - 1) / 2) * bar_width
            color = s['colors'][gi]
            is_hollow = s.get('bar_fill', 'Solid') == 'Hollow'
            edge_w = s.get('bar_edge_width', 1.5)
            if is_hollow:
                ax.bar(x + offset, means, bar_width * 0.85,
                       facecolor='none', edgecolor=color, linewidth=edge_w, label=group['name'])
            else:
                ax.bar(x + offset, means, bar_width * 0.85,
                       color=color, alpha=0.7, edgecolor=color, linewidth=0.5, label=group['name'])
            ax.errorbar(x + offset, means, yerr=errs, fmt='none', ecolor='black',
                        capsize=capsize, elinewidth=s['error_linewidth'],
                        capthick=s['error_linewidth'])
            for pi, vals in enumerate(all_vals):
                scatter_points(ax, x[pi] + offset, vals, color, s)

        # Statistics
        if stat_test and n_groups >= 2:
            group_names = [g['name'] for g in groups]

            if _is_twoway(stat_test):
                # ── Two-way ANOVA: build long-format across all periods ──
                rows_long = []
                for gi, group in enumerate(groups):
                    for pname in period_names:
                        vals = compute_total_time(time_data, group['ids'], periods[pname])[state]
                        for mi, v in enumerate(vals):
                            rows_long.append({
                                'value': float(v),
                                'group': group['name'],
                                'period': pname,
                                'subject': f"{group['name']}_{group['ids'][mi]}",
                            })
                long_df = pd.DataFrame(rows_long)
                export_store = {}
                twoway_results = run_twoway_anova(long_df, dv='value',
                                                  between='group', within='period',
                                                  test_type=stat_test, posthoc=posthoc,
                                                  _export_store=export_store)
                # Store for Excel export (keyed by state)
                if export_store:
                    if '_twoway_exports' not in st.session_state:
                        st.session_state._twoway_exports = {}
                    st.session_state._twoway_exports[state] = export_store
                alpha = s.get('sig_alpha', 0.05)
                for r in twoway_results:
                    r['State'] = state
                    # Ensure all rows have p_corrected and Significance
                    if 'p_corrected' not in r:
                        r['p_corrected'] = r.get('p', np.nan)
                    if 'Significance' not in r:
                        p_c = r.get('p_corrected', r.get('p', np.nan))
                        r['Significance'] = p_to_stars(p_c, alpha) if not np.isnan(p_c) else '-'
                stat_results_all.extend(twoway_results)

                # Draw brackets for pairwise post-hoc results
                # Šídák post-hoc results already have adjusted p and row_level
                pairwise = [r for r in twoway_results if 'g1' in r and 'g2' in r]
                if pairwise:
                    for r in pairwise:
                        if 'Significance' not in r:
                            r['Significance'] = p_to_stars(r.get('p', np.nan), alpha)

                    # Draw per-period brackets matching row_level
                    for pi_idx, pname in enumerate(period_names):
                        gv = all_group_data[pi_idx]
                        bracket_y_base = max(np.max(v) for v in gv if len(v) > 0)
                        bracket_count = 0
                        # Filter pairwise results for this period
                        period_pairs = [r for r in pairwise
                                        if r.get('row_level', '') == pname
                                        or (r.get('row_level', '') == '' and pi_idx == 0)]
                        for r in period_pairs:
                            if r['g1'] not in group_names or r['g2'] not in group_names:
                                continue
                            p = r.get('p', np.nan)
                            try:
                                p = float(p)
                            except (TypeError, ValueError):
                                continue
                            if np.isnan(p):
                                continue
                            label = format_p(p, s.get('sig_display', 'Stars'),
                                             s.get('show_ns', False), alpha)
                            if label is None:
                                continue
                            gi1 = group_names.index(r['g1'])
                            gi2 = group_names.index(r['g2'])
                            x1 = pi_idx + (gi1 - (n_groups - 1) / 2) * bar_width
                            x2 = pi_idx + (gi2 - (n_groups - 1) / 2) * bar_width
                            y_offset = bracket_y_base * (1 + s['sig_height'] * (1 + bracket_count * 1.5))
                            add_significance_bracket(ax, x1, x2, y_offset, label,
                                                     lw=s['sig_line_width'], fs=s['sig_font_size'])
                            bracket_count += 1
            else:
                # ── One-factor tests: per-period pairwise ──
                for pi_idx in range(n_periods):
                    gv = all_group_data[pi_idx]
                    results = run_pairwise(gv, stat_test, group_names, posthoc=posthoc)

                    pairwise = [r for r in results if 'g1' in r and 'g2' in r]

                    if pairwise:
                        p_vals = [r['p'] for r in pairwise]
                        p_corrected = apply_correction(p_vals, s.get('stat_correction', 'None'))
                        for ri, r in enumerate(pairwise):
                            r['p_corrected'] = p_corrected[ri]
                            r['Significance'] = p_to_stars(p_corrected[ri], s.get('sig_alpha', 0.05))

                    for r in results:
                        r['State'] = state
                        r['Period'] = period_names[pi_idx]
                    stat_results_all.extend(results)

                    bracket_y_base = max(np.max(v) for v in gv if len(v) > 0)
                    for ri, r in enumerate(pairwise):
                        p = r.get('p_corrected', r['p'])
                        if np.isnan(p):
                            continue
                        label = format_p(p, s.get('sig_display', 'Stars'), s.get('show_ns', False),
                                         s.get('sig_alpha', 0.05))
                        if label is None:
                            continue
                        gi1 = group_names.index(r['g1'])
                        gi2 = group_names.index(r['g2'])
                        x1 = pi_idx + (gi1 - (n_groups - 1) / 2) * bar_width
                        x2 = pi_idx + (gi2 - (n_groups - 1) / 2) * bar_width
                        y_offset = bracket_y_base * (1 + s['sig_height'] * (1 + ri * 1.5))
                        add_significance_bracket(ax, x1, x2, y_offset, label,
                                                 lw=s['sig_line_width'], fs=s['sig_font_size'])

        ax.set_xticks(np.arange(n_periods))
        ax.set_xticklabels(period_names, fontsize=fs)
        ax.set_title(state, fontsize=fs + 2)
        apply_axis_settings(ax, s, default_ylabel='Total time (min)')
        if si == 0:
            ax.legend(fontsize=fs - 2, frameon=False)

    fig.tight_layout()
    return fig, stat_results_all


# ─── Plot: Time Course ──────────────────────────────────────────────────────

def plot_timecourse(parsed, groups, s):
    time_data = parsed['Time']
    fs = s['font_size']
    stat_test = s.get('active_stat')
    stat_results_all = []
    fig, axes = plt.subplots(1, 3, figsize=(s['fig_width'], s['fig_height']))
    fig.subplots_adjust(wspace=s['wspace'])

    for si, state in enumerate(STATES):
        ax = axes[si]
        group_means, group_sems = [], []
        for gi, group in enumerate(groups):
            avg_df = average_across_basals(time_data, state, group['ids'])
            mean, sem = group_mean_sem(avg_df, error_type=s['error_type'])
            x = mean.index.values
            color = s['colors'][gi]
            shade_color = s['shade_color_list'][gi]
            ax.plot(x, mean, color=color, label=group['name'],
                    linewidth=s['line_width'], linestyle=s['linestyle'])
            draw_error(ax, x, mean, sem, color, shade_color, s, gi)
            group_means.append(mean)
            group_sems.append(sem)

        if stat_test and len(groups) >= 2:
            x_vals = group_means[0].index.values
            group_names = [g['name'] for g in groups]

            if _is_twoway(stat_test):
                # Build long-format: value, group, timepoint, subject
                rows_long = []
                for gi, group in enumerate(groups):
                    avg_df = average_across_basals(time_data, state, group['ids'])
                    for xi in x_vals:
                        if xi not in avg_df.index:
                            continue
                        for mi, mid in enumerate(group['ids']):
                            v = avg_df.loc[xi, mid] if mid in avg_df.columns else np.nan
                            rows_long.append({
                                'value': float(v), 'group': group['name'],
                                'timepoint': str(int(xi)),
                                'subject': f"{group['name']}_{mid}",
                            })
                long_df = pd.DataFrame(rows_long).dropna(subset=['value'])
                twoway_results = run_twoway_anova(long_df, dv='value',
                                                  between='group', within='timepoint',
                                                  test_type=stat_test,
                                                  posthoc=s.get('stat_posthoc', 'Sidak'))
                alpha = s.get('sig_alpha', 0.05)
                for r in twoway_results:
                    r['State'] = state
                    if 'p_corrected' not in r:
                        r['p_corrected'] = r.get('p', np.nan)
                    if 'Significance' not in r:
                        p_c = r.get('p_corrected', r.get('p', np.nan))
                        r['Significance'] = p_to_stars(p_c, alpha) if not np.isnan(p_c) else '-'
                stat_results_all.extend(twoway_results)

            # Pointwise markers (for both 1F and 2F: shows per-timepoint significance)
            p_vals_raw, y_tops = [], []
            for xi in x_vals:
                dl = []
                for group in groups:
                    avg_df = average_across_basals(time_data, state, group['ids'])
                    dl.append(avg_df.loc[xi].values if xi in avg_df.index else np.array([]))
                p_vals_raw.append(run_pointwise_test(dl, stat_test, group_names))
                y_tops.append(max(gm.get(xi, 0) + gs.get(xi, 0)
                                  for gm, gs in zip(group_means, group_sems)) * (1 + s['sig_height']))

            p_corrected = apply_correction(p_vals_raw, s.get('stat_correction', 'None'))
            add_significance_markers(ax, x_vals, p_corrected, y_tops, s)

            if not _is_twoway(stat_test):
                for xi, p_raw, p_cor in zip(x_vals, p_vals_raw, p_corrected):
                    sig = p_to_stars(p_cor, s.get('sig_alpha', 0.05))
                    stat_results_all.append({
                        'State': state, 'Timepoint': int(xi), 'Test': stat_test,
                        'p': p_raw, 'p_corrected': p_cor, 'Significance': sig,
                    })

        ax.set_title(state, fontsize=fs + 2)
        add_light_dark_shading(ax, s)
        if s.get('x_tick', 0) == 0:
            ax.set_xticks(range(0, 24, 2))
        apply_axis_settings(ax, s, default_xlabel='ZT (h)', default_ylabel='Time (min/h)')
        if si == 0:
            ax.legend(fontsize=fs - 2, frameon=False)

    fig.tight_layout()
    return fig, stat_results_all


# ─── Plot: Delta Curves ─────────────────────────────────────────────────────

def plot_delta_curve(parsed, groups, sheet_name, default_ylabel, state, s):
    sheet_data = parsed[sheet_name]
    fs = s['font_size']
    stat_test = s.get('active_stat')
    stat_results_all = []
    fig, ax = plt.subplots(figsize=(s['fig_width'], s['fig_height']))

    group_means, group_sems = [], []
    for gi, group in enumerate(groups):
        avg_df = average_across_basals(sheet_data, state, group['ids'], exclude_zero=True)
        mean, sem = group_mean_sem(avg_df, error_type=s['error_type'])
        x = mean.index.values
        color = s['colors'][gi]
        shade_color = s['shade_color_list'][gi]
        ax.plot(x, mean, color=color, label=group['name'],
                linewidth=s['line_width'], linestyle=s['linestyle'])
        draw_error(ax, x, mean, sem, color, shade_color, s, gi)
        group_means.append(mean)
        group_sems.append(sem)

    if stat_test and len(groups) >= 2:
        x_vals = group_means[0].index.values
        group_names = [g['name'] for g in groups]

        if _is_twoway(stat_test):
            rows_long = []
            for gi, group in enumerate(groups):
                avg_df = average_across_basals(sheet_data, state, group['ids'], exclude_zero=True)
                for xi in x_vals:
                    if xi not in avg_df.index:
                        continue
                    for mi, mid in enumerate(group['ids']):
                        v = avg_df.loc[xi, mid] if mid in avg_df.columns else np.nan
                        if not np.isnan(v):
                            rows_long.append({
                                'value': float(v), 'group': group['name'],
                                'timepoint': str(int(xi)),
                                'subject': f"{group['name']}_{mid}",
                            })
            long_df = pd.DataFrame(rows_long)
            if not long_df.empty:
                twoway_results = run_twoway_anova(long_df, dv='value',
                                                  between='group', within='timepoint',
                                                  test_type=stat_test,
                                                  posthoc=s.get('stat_posthoc', 'Sidak'))
                alpha = s.get('sig_alpha', 0.05)
                for r in twoway_results:
                    if 'p_corrected' not in r:
                        r['p_corrected'] = r.get('p', np.nan)
                    if 'Significance' not in r:
                        p_c = r.get('p_corrected', r.get('p', np.nan))
                        r['Significance'] = p_to_stars(p_c, alpha) if not np.isnan(p_c) else '-'
                stat_results_all.extend(twoway_results)

        # Pointwise markers
        p_vals_raw, y_tops = [], []
        for xi in x_vals:
            dl = []
            for group in groups:
                avg_df = average_across_basals(sheet_data, state, group['ids'], exclude_zero=True)
                if xi in avg_df.index:
                    v = avg_df.loc[xi].values
                    dl.append(v[~np.isnan(v)])
                else:
                    dl.append(np.array([]))
            p_vals_raw.append(run_pointwise_test(dl, stat_test, group_names))
            y_tops.append(max(gm.get(xi, 0) + gs.get(xi, 0)
                              for gm, gs in zip(group_means, group_sems)) * (1 + s['sig_height']))

        p_corrected = apply_correction(p_vals_raw, s.get('stat_correction', 'None'))
        add_significance_markers(ax, x_vals, p_corrected, y_tops, s)

        if not _is_twoway(stat_test):
            for xi, p_raw, p_cor in zip(x_vals, p_vals_raw, p_corrected):
                sig = p_to_stars(p_cor, s.get('sig_alpha', 0.05))
                stat_results_all.append({
                    'Timepoint': int(xi), 'Test': stat_test,
                    'p': p_raw, 'p_corrected': p_cor, 'Significance': sig,
                })

    ax.set_title(f'{sheet_name} ({state})', fontsize=fs + 2)
    if s.get('x_tick', 0) == 0:
        ax.set_xticks(range(0, 24, 2))
    add_light_dark_shading(ax, s)
    apply_axis_settings(ax, s, default_xlabel='ZT (h)', default_ylabel=default_ylabel)
    ax.legend(fontsize=fs - 2, frameon=False)
    fig.tight_layout()
    return fig, stat_results_all


# ─── Plot: Power Spectrum ───────────────────────────────────────────────────

def plot_spectrum(parsed, groups, state, s):
    spec_sheets = ['Spectrum-24h', 'Spectrum-light', 'Spectrum-dark']
    titles = ['24h', 'Light', 'Dark']
    fs = s['font_size']
    stat_test = s.get('active_stat')
    stat_results_all = []
    fig, axes = plt.subplots(1, 3, figsize=(s['fig_width'], s['fig_height']))
    fig.subplots_adjust(wspace=s['wspace'])

    for pi, (sheet_name, title) in enumerate(zip(spec_sheets, titles)):
        ax = axes[pi]
        sheet_data = parsed[sheet_name]
        group_means, group_sems = [], []
        for gi, group in enumerate(groups):
            avg_df = average_across_basals(sheet_data, state, group['ids'])
            mean, sem = group_mean_sem(avg_df, error_type=s['error_type'])
            x = mean.index.values
            color = s['colors'][gi]
            shade_color = s['shade_color_list'][gi]
            ax.plot(x, mean, color=color, label=group['name'],
                    linewidth=s['line_width'], linestyle=s['linestyle'])
            draw_error(ax, x, mean, sem, color, shade_color, s, gi)
            group_means.append(mean)
            group_sems.append(sem)

        if stat_test and len(groups) >= 2:
            x_vals = group_means[0].index.values
            group_names = [g['name'] for g in groups]

            if _is_twoway(stat_test):
                rows_long = []
                for gi, group in enumerate(groups):
                    avg_df = average_across_basals(sheet_data, state, group['ids'])
                    for xi in x_vals:
                        if xi not in avg_df.index:
                            continue
                        for mi, mid in enumerate(group['ids']):
                            v = avg_df.loc[xi, mid] if mid in avg_df.columns else np.nan
                            if not pd.isna(v):
                                rows_long.append({
                                    'value': float(v), 'group': group['name'],
                                    'frequency': str(int(xi)),
                                    'subject': f"{group['name']}_{mid}",
                                })
                long_df = pd.DataFrame(rows_long)
                if not long_df.empty:
                    twoway_results = run_twoway_anova(long_df, dv='value',
                                                      between='group', within='frequency',
                                                      test_type=stat_test,
                                                      posthoc=s.get('stat_posthoc', 'Sidak'))
                    alpha = s.get('sig_alpha', 0.05)
                    for r in twoway_results:
                        r['Period'] = title
                        if 'p_corrected' not in r:
                            r['p_corrected'] = r.get('p', np.nan)
                        if 'Significance' not in r:
                            p_c = r.get('p_corrected', r.get('p', np.nan))
                            r['Significance'] = p_to_stars(p_c, alpha) if not np.isnan(p_c) else '-'
                    stat_results_all.extend(twoway_results)

            # Pointwise markers
            p_vals_raw, y_tops = [], []
            for xi in x_vals:
                dl = []
                for group in groups:
                    avg_df = average_across_basals(sheet_data, state, group['ids'])
                    dl.append(avg_df.loc[xi].values if xi in avg_df.index else np.array([]))
                p_vals_raw.append(run_pointwise_test(dl, stat_test, group_names))
                y_tops.append(max(gm.get(xi, 0) + gs.get(xi, 0)
                                  for gm, gs in zip(group_means, group_sems)) * (1 + s['sig_height']))

            p_corrected = apply_correction(p_vals_raw, s.get('stat_correction', 'None'))
            add_significance_markers(ax, x_vals, p_corrected, y_tops, s)

            if not _is_twoway(stat_test):
                for xi, p_raw, p_cor in zip(x_vals, p_vals_raw, p_corrected):
                    sig = p_to_stars(p_cor, s.get('sig_alpha', 0.05))
                    stat_results_all.append({
                        'Period': title, 'Frequency': int(xi), 'Test': stat_test,
                        'p': p_raw, 'p_corrected': p_cor, 'Significance': sig,
                    })

        ax.set_title(f'{title} ({state})', fontsize=fs + 2)
        apply_axis_settings(ax, s, default_xlabel='Frequency (Hz)',
                            default_ylabel='Normalized power' if pi == 0 else '')
        if pi == 0:
            ax.legend(fontsize=fs - 2, frameon=False)

    fig.tight_layout()
    return fig, stat_results_all


# ─── Statistics Results Table ────────────────────────────────────────────────

def _fmt_num(x, fmt='.4f', na_str='N/A'):
    """Format a numeric value robustly (handles numpy scalars, Python floats, NaN)."""
    try:
        f = float(x)
        if pd.isna(f) or np.isnan(f):
            return na_str
        return f'{f:{fmt}}'
    except (TypeError, ValueError):
        return na_str


def show_stat_results_table(stat_results, key_prefix):
    if not stat_results:
        return
    df = pd.DataFrame(stat_results)
    # Format numeric columns robustly
    for col in ['p', 'p_raw', 'p_corrected']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: _fmt_num(x, '.6f', 'N/A'))
    for col in ['Statistic', 'Mean Diff', 'CI_low', 'CI_high', 'SS', 'MS']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: _fmt_num(x, '.4f', '-'))
    if 'Effect size' in df.columns:
        df['Effect size'] = df['Effect size'].apply(lambda x: _fmt_num(x, '.4f', '-'))
    # Remove internal keys
    display_cols = [c for c in df.columns if c not in ('g1', 'g2', 'row_level')]
    # Reorder: put key columns first
    preferred = ['Comparison', 'Test', 'SS', 'df', 'MS', 'Statistic', 'Mean Diff',
                 'p_raw', 'p', 'p_corrected', 'CI_low', 'CI_high',
                 'Effect size', 'Significance']
    ordered = [c for c in preferred if c in display_cols] + [c for c in display_cols if c not in preferred]
    with st.expander("Statistical Results", expanded=False):
        st.dataframe(df[ordered], use_container_width=True, hide_index=True)


# ─── Excel Export of Statistics ──────────────────────────────────────────────

def export_stats_to_excel(long_df, dv, row_factor, col_factor, aov_result, posthoc_results):
    """Export all statistical results to an Excel file (5 sheets) as BytesIO."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    # ── Sheet 1: Raw Data ──
    ws1 = wb.active
    ws1.title = "Raw Data"
    headers1 = ['Subject', 'Row_Factor', 'Column_Factor', 'Value']
    for ci, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=ci, value=h).font = bold
    for ri, (_, row) in enumerate(long_df.iterrows(), 2):
        ws1.cell(row=ri, column=1, value=str(row.get('subject', '')))
        ws1.cell(row=ri, column=2, value=str(row[row_factor]))
        ws1.cell(row=ri, column=3, value=str(row[col_factor]))
        ws1.cell(row=ri, column=4, value=float(row[dv]) if not pd.isna(row[dv]) else None)

    # ── Sheet 2: Group Means ──
    ws2 = wb.create_sheet("Group Means")
    headers2 = ['Row_Factor', 'Column_Factor', 'N', 'Mean', 'SD', 'SEM']
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=ci, value=h).font = bold
    ri = 2
    for row_lev in sorted(long_df[row_factor].unique()):
        for col_lev in sorted(long_df[col_factor].unique()):
            vals = long_df[(long_df[row_factor] == row_lev) &
                           (long_df[col_factor] == col_lev)][dv].dropna().values
            n = len(vals)
            mean = np.mean(vals) if n > 0 else None
            sd = float(np.std(vals, ddof=1)) if n > 1 else None
            sem = sd / np.sqrt(n) if sd is not None and n > 0 else None
            ws2.cell(row=ri, column=1, value=str(row_lev))
            ws2.cell(row=ri, column=2, value=str(col_lev))
            ws2.cell(row=ri, column=3, value=n)
            ws2.cell(row=ri, column=4, value=mean)
            ws2.cell(row=ri, column=5, value=sd)
            ws2.cell(row=ri, column=6, value=sem)
            ri += 1

    # ── Sheet 3: ANOVA Table ──
    ws3 = wb.create_sheet("ANOVA Table")
    headers3 = ['Source', 'SS', 'DF', 'MS', 'F', 'P_value']
    for ci, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=ci, value=h).font = bold
    for ri, row in enumerate(aov_result['table'], 2):
        ws3.cell(row=ri, column=1, value=row['Source'])
        ws3.cell(row=ri, column=2, value=row['SS'])
        ws3.cell(row=ri, column=3, value=row['DF'])
        ws3.cell(row=ri, column=4, value=row['MS'])
        f_val = row['F']
        ws3.cell(row=ri, column=5, value=f_val if not np.isnan(f_val) else None)
        p_val = row['p']
        ws3.cell(row=ri, column=6, value=p_val if not np.isnan(p_val) else None)

    # ── Sheet 4: Sidak Comparisons ──
    ws4 = wb.create_sheet("Sidak Comparisons")
    headers4 = ['Family', 'Comparison', 'Mean1', 'Mean2', 'Mean_Diff',
                'MSE_used', 'DFe_used', 'n1', 'n2', 'SE_of_diff',
                't_statistic', 'p_raw', 'k', 'p_adjusted', 'Significance']
    for ci, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=ci, value=h).font = bold
    MSE = aov_result['MSE']
    DFe = aov_result['DFe']
    col_levels = sorted(long_df[col_factor].unique())
    k = len(col_levels) * (len(col_levels) - 1) // 2

    ri = 2
    for r in posthoc_results:
        row_lev = r.get('row_level', r.get('Comparison', '').split(']')[0].replace('[', ''))
        comp = f"{r['g1']} vs {r['g2']}"
        # Get individual group means
        g1_vals = long_df[(long_df[row_factor] == row_lev) &
                          (long_df[col_factor] == r['g1'])][dv].dropna().values
        g2_vals = long_df[(long_df[row_factor] == row_lev) &
                          (long_df[col_factor] == r['g2'])][dv].dropna().values
        n1, n2 = len(g1_vals), len(g2_vals)
        mean1 = np.mean(g1_vals) if n1 > 0 else None
        mean2 = np.mean(g2_vals) if n2 > 0 else None
        se = np.sqrt(MSE * (1.0/n1 + 1.0/n2)) if n1 > 0 and n2 > 0 else None

        ws4.cell(row=ri, column=1, value=str(row_lev))
        ws4.cell(row=ri, column=2, value=comp)
        ws4.cell(row=ri, column=3, value=mean1)
        ws4.cell(row=ri, column=4, value=mean2)
        ws4.cell(row=ri, column=5, value=_safe_float(r.get('Mean Diff', np.nan)))
        ws4.cell(row=ri, column=6, value=MSE)
        ws4.cell(row=ri, column=7, value=DFe)
        ws4.cell(row=ri, column=8, value=n1)
        ws4.cell(row=ri, column=9, value=n2)
        ws4.cell(row=ri, column=10, value=se)
        ws4.cell(row=ri, column=11, value=_safe_float(r.get('Statistic', np.nan)))
        ws4.cell(row=ri, column=12, value=_safe_float(r.get('p_raw', np.nan)))
        ws4.cell(row=ri, column=13, value=k)
        ws4.cell(row=ri, column=14, value=_safe_float(r.get('p', np.nan)))
        ws4.cell(row=ri, column=15, value=r.get('Significance', '-'))
        ri += 1

    # ── Sheet 5: Verification ──
    ws5 = wb.create_sheet("Verification")
    for ci, h in enumerate(['Parameter', 'Value'], 1):
        ws5.cell(row=1, column=ci, value=h).font = bold
    verification = [
        ('MSE (Residual Mean Square)', MSE),
        ('DFe (Residual df)', DFe),
        ('Sidak k (comparisons per family)', k),
        ('Number of row levels', len(long_df[row_factor].unique())),
        ('Number of column levels', len(col_levels)),
        ('Total N', len(long_df)),
        ('Total SS', sum(r['SS'] for r in aov_result['table'])),
        ('Total DF', sum(r['DF'] for r in aov_result['table'])),
    ]
    # Add per-group N
    for col_lev in sorted(long_df[col_factor].unique()):
        n_col = len(long_df[long_df[col_factor] == col_lev])
        verification.append((f'N ({col_lev})', n_col))
    for row_lev in sorted(long_df[row_factor].unique()):
        for col_lev in sorted(long_df[col_factor].unique()):
            n_cell = len(long_df[(long_df[row_factor] == row_lev) &
                                 (long_df[col_factor] == col_lev)])
            verification.append((f'N ({row_lev}, {col_lev})', n_cell))

    for ri, (param, val) in enumerate(verification, 2):
        ws5.cell(row=ri, column=1, value=param)
        ws5.cell(row=ri, column=2, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Render a single plot panel ─────────────────────────────────────────────

def render_plot_panel(plot_key, plot_type, parsed, groups, delta_state=None, spec_state=None):
    n_groups = len(groups)

    if plot_key == 'bar':
        settings = get_plot_settings("bar", groups, plot_type='bar', n_groups=n_groups)
        fig, stat_results = plot_bar_chart(parsed, groups, settings)
    elif plot_key == 'tc':
        settings = get_plot_settings("tc", groups, plot_type='line', n_groups=n_groups)
        fig, stat_results = plot_timecourse(parsed, groups, settings)
    elif plot_key == 'delta':
        settings = get_plot_settings("delta", groups, plot_type='line', n_groups=n_groups)
        fig, stat_results = plot_delta_curve(parsed, groups, 'delta power density',
                                             'Delta power density', state=delta_state, s=settings)
    elif plot_key == 'pct':
        settings = get_plot_settings("pct", groups, plot_type='line', n_groups=n_groups)
        fig, stat_results = plot_delta_curve(parsed, groups, 'percentage',
                                             'Delta power (%)', state=delta_state, s=settings)
    elif plot_key == 'spec':
        settings = get_plot_settings("spec", groups, plot_type='spectrum', n_groups=n_groups)
        fig, stat_results = plot_spectrum(parsed, groups, state=spec_state, s=settings)
    else:
        return

    st.pyplot(fig, use_container_width=False, transparent=False)
    show_export_buttons(fig, plot_key)
    if stat_results:
        show_stat_results_table(stat_results, plot_key)

        # Excel export for Two-way ANOVA results
        twoway_exports = st.session_state.get('_twoway_exports', {})
        if twoway_exports:
            # Use first available state's export data
            for state_key, export_data in twoway_exports.items():
                if all(k in export_data for k in ('aov_result', 'posthoc_results', 'long_df')):
                    try:
                        xlsx_buf = export_stats_to_excel(
                            export_data['long_df'], export_data['dv'],
                            export_data['row_factor'], export_data['col_factor'],
                            export_data['aov_result'], export_data['posthoc_results'])
                        st.download_button(
                            f"Stats Excel ({state_key})",
                            xlsx_buf,
                            f"stats_results_{state_key}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"{plot_key}_xlsx_{state_key}")
                    except Exception as e:
                        st.caption(f"Excel export error: {e}")

    plt.close(fig)


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(layout="wide", page_title="SomnoCore AI — Sleep EEG")

    # ═══════════════════════════════════════════════════════════════════════════
    # Linear / Notion–style CSS overhaul  (compact + Inter)
    # ═══════════════════════════════════════════════════════════════════════════
    st.markdown("""
    <style>
    /* ── Import Inter font ─────────────────────────────────────────────────── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600&display=swap');

    /* ── Hide Streamlit chrome ─────────────────────────────────────────────── */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    [data-testid="stToolbar"] {display: none !important;}
    .stDeployButton {display: none !important;}

    /* ── Global typography & background ────────────────────────────────────── */
    /* Target text-bearing elements only — do NOT use * to avoid breaking icon fonts */
    html, body, [data-testid="stAppViewContainer"],
    [data-testid="stAppViewContainer"] p,
    [data-testid="stAppViewContainer"] span:not([data-testid="stIconMaterial"]):not(.material-symbols-rounded):not(.material-icons),
    [data-testid="stAppViewContainer"] div,
    [data-testid="stAppViewContainer"] input,
    [data-testid="stAppViewContainer"] button,
    [data-testid="stAppViewContainer"] select,
    [data-testid="stAppViewContainer"] textarea,
    [data-testid="stAppViewContainer"] label,
    [data-testid="stAppViewContainer"] h1,
    [data-testid="stAppViewContainer"] h2,
    [data-testid="stAppViewContainer"] h3,
    [data-testid="stAppViewContainer"] h4,
    [data-testid="stAppViewContainer"] h5,
    [data-testid="stAppViewContainer"] h6,
    [data-testid="stAppViewContainer"] td,
    [data-testid="stAppViewContainer"] th,
    [data-testid="stAppViewContainer"] li,
    [data-testid="stAppViewContainer"] a {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }
    /* Restore icon fonts explicitly */
    .material-symbols-rounded, .material-icons,
    [data-testid="stIconMaterial"],
    [data-testid="stExpander"] summary svg,
    [data-testid="stExpander"] details summary span[class*="icon"],
    [data-testid="stExpander"] details summary span[class*="Icon"] {
        font-family: 'Material Symbols Rounded', 'Material Icons', sans-serif !important;
    }
    html, body, [data-testid="stAppViewContainer"] {
        background-color: #F8FAFC !important;
        color: #0F172A !important;
        line-height: 1.6;
    }
    [data-testid="stAppViewContainer"] > .main {
        background-color: #F8FAFC !important;
    }
    .main .block-container {
        padding-top: 1.2rem !important;
        padding-bottom: 1.2rem !important;
        max-width: 1200px;
    }

    /* ── Headings ──────────────────────────────────────────────────────────── */
    h1, h2, h3, h4, h5, h6,
    [data-testid="stHeadingWithActionElements"] {
        font-weight: 600 !important;
        letter-spacing: -0.02em !important;
        color: #0F172A !important;
    }
    h1 { font-size: 1.5rem !important; margin-bottom: 0.3rem !important; }
    h2 { font-size: 1.15rem !important; margin-bottom: 0.2rem !important; }
    h3 { font-size: 1rem !important; margin-bottom: 0.15rem !important; }

    /* ── Section label / caption: Notion style ─────────────────────────────── */
    .stCaption, [data-testid="stCaptionContainer"],
    .stMarkdown small, .element-container .stCaption p {
        color: #94A3B8 !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        letter-spacing: 0.08em !important;
        text-transform: uppercase !important;
        margin-bottom: 2px !important;
        line-height: 1.3 !important;
    }

    /* ── Sidebar ──────────────────────────────────────────────────────────── */
    [data-testid="stSidebar"] {
        background-color: #F1F5F9 !important;
        border-right: 1px solid #E2E8F0 !important;
    }
    [data-testid="stSidebar"] > div:first-child {
        background-color: #F1F5F9 !important;
        padding-top: 1rem !important;
    }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        color: #0F172A !important;
    }
    [data-testid="stSidebar"] .stMarkdown p {
        color: #334155 !important;
        font-size: 13px !important;
    }

    /* ── Cards / Expanders (compact) ───────────────────────────────────────── */
    [data-testid="stExpander"] {
        background-color: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 10px !important;
        box-shadow: none !important;
        margin-bottom: 8px !important;
        overflow: hidden;
    }
    [data-testid="stExpander"] details {
        border: none !important;
    }
    [data-testid="stExpander"] summary {
        font-weight: 500 !important;
        font-size: 13px !important;
        color: #0F172A !important;
        padding: 8px 12px !important;
        border-radius: 10px !important;
        min-height: unset !important;
    }
    [data-testid="stExpander"] summary:hover {
        background-color: #F8FAFC !important;
    }
    [data-testid="stExpander"] [data-testid="stExpanderDetails"] {
        padding: 0 12px 12px 12px !important;
    }

    /* ── Buttons: Primary (Generate Plots) ─────────────────────────────────── */
    .stButton > button[kind="primary"],
    .stButton > button[data-testid="stBaseButton-primary"] {
        background-color: #2563EB !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        padding: 0 20px !important;
        height: 48px !important;
        min-height: 48px !important;
        width: 100% !important;
        transition: all 0.15s ease !important;
        box-shadow: 0 1px 2px rgba(37, 99, 235, 0.15) !important;
        letter-spacing: -0.01em !important;
    }
    .stButton > button[kind="primary"]:hover,
    .stButton > button[data-testid="stBaseButton-primary"]:hover {
        background-color: #1D4ED8 !important;
        box-shadow: 0 2px 8px rgba(37, 99, 235, 0.25) !important;
        transform: translateY(-1px);
    }
    .stButton > button[kind="primary"]:active,
    .stButton > button[data-testid="stBaseButton-primary"]:active {
        background-color: #1E40AF !important;
        transform: translateY(0);
    }

    /* ── Buttons: Secondary / default ──────────────────────────────────────── */
    .stButton > button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]) {
        background-color: #FFFFFF !important;
        color: #334155 !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 6px !important;
        font-weight: 500 !important;
        font-size: 13px !important;
        transition: all 0.15s ease !important;
        padding: 4px 10px !important;
        min-height: 30px !important;
        height: 30px !important;
    }
    .stButton > button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]):hover {
        background-color: #F8FAFC !important;
        border-color: #CBD5E1 !important;
    }

    /* ── Download buttons: outline style ───────────────────────────────────── */
    .stDownloadButton > button {
        background-color: #FFFFFF !important;
        color: #2563EB !important;
        border: 1.5px solid #2563EB !important;
        border-radius: 6px !important;
        font-weight: 500 !important;
        font-size: 13px !important;
        transition: all 0.15s ease !important;
        height: 34px !important;
        min-height: 34px !important;
        padding: 0 14px !important;
    }
    .stDownloadButton > button:hover {
        background-color: #EFF6FF !important;
        border-color: #1D4ED8 !important;
        color: #1D4ED8 !important;
    }

    /* ── Inputs: compact text, number, selectbox ───────────────────────────── */
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input,
    .stTextInput input,
    .stNumberInput input {
        border: 1px solid #E2E8F0 !important;
        border-radius: 6px !important;
        background-color: #FFFFFF !important;
        color: #0F172A !important;
        font-size: 13px !important;
        padding: 4px 8px !important;
        height: 32px !important;
        min-height: 32px !important;
        transition: border-color 0.15s ease !important;
    }
    [data-testid="stTextInput"] input:focus,
    [data-testid="stNumberInput"] input:focus,
    .stTextInput input:focus,
    .stNumberInput input:focus {
        border-color: #2563EB !important;
        box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.08) !important;
        outline: none !important;
    }

    /* Selectbox (compact) */
    [data-testid="stSelectbox"] > div > div {
        border: 1px solid #E2E8F0 !important;
        border-radius: 6px !important;
        background-color: #FFFFFF !important;
        min-height: 32px !important;
        font-size: 13px !important;
    }
    [data-testid="stSelectbox"] > div > div:focus-within {
        border-color: #2563EB !important;
        box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.08) !important;
    }

    /* Multiselect (compact) */
    [data-testid="stMultiSelect"] > div > div {
        border: 1px solid #E2E8F0 !important;
        border-radius: 6px !important;
        background-color: #FFFFFF !important;
        min-height: 32px !important;
        font-size: 13px !important;
    }
    [data-testid="stMultiSelect"] > div > div:focus-within {
        border-color: #2563EB !important;
        box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.08) !important;
    }

    /* ── Number input stepper buttons (compact) ────────────────────────────── */
    [data-testid="stNumberInput"] button {
        border: 1px solid #E2E8F0 !important;
        background-color: #F8FAFC !important;
        color: #64748B !important;
        height: 16px !important;
        width: 24px !important;
    }
    [data-testid="stNumberInput"] button:hover {
        background-color: #EFF6FF !important;
        border-color: #2563EB !important;
        color: #2563EB !important;
    }
    /* Make number_input container tighter */
    [data-testid="stNumberInput"] > div {
        gap: 0 !important;
    }

    /* ── Tabs: underline style ─────────────────────────────────────────────── */
    [data-testid="stTabs"] {
        background: transparent !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        gap: 0 !important;
        border-bottom: 1px solid #E2E8F0 !important;
        background: transparent !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab"] {
        background-color: transparent !important;
        border: none !important;
        border-bottom: 2px solid transparent !important;
        border-radius: 0 !important;
        color: #64748B !important;
        font-weight: 500 !important;
        font-size: 13px !important;
        padding: 6px 12px !important;
        margin-bottom: -1px !important;
        transition: all 0.15s ease !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab"]:hover {
        color: #0F172A !important;
        background-color: #F8FAFC !important;
    }
    [data-testid="stTabs"] [aria-selected="true"] {
        color: #2563EB !important;
        border-bottom: 2px solid #2563EB !important;
        background-color: transparent !important;
        font-weight: 600 !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-highlight"] {
        display: none !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-border"] {
        display: none !important;
    }

    /* ── Checkbox / Radio (compact) ────────────────────────────────────────── */
    [data-testid="stCheckbox"] input[type="checkbox"] {
        accent-color: #2563EB !important;
    }
    [data-testid="stCheckbox"] label span {
        color: #0F172A !important;
        font-size: 13px !important;
    }
    [data-testid="stCheckbox"] label {
        padding: 0 !important;
        min-height: unset !important;
    }
    [data-testid="stRadio"] input[type="radio"],
    .stRadio input[type="radio"] {
        accent-color: #2563EB !important;
    }
    [data-testid="stRadio"] label,
    .stRadio label {
        color: #334155 !important;
        font-size: 13px !important;
    }
    [data-testid="stRadio"] [data-baseweb="radio"] label > div:first-child {
        border-color: #CBD5E1 !important;
    }

    /* ── File uploader (compact) ───────────────────────────────────────────── */
    [data-testid="stFileUploader"] {
        background-color: #FFFFFF !important;
        border: 1.5px dashed #E2E8F0 !important;
        border-radius: 10px !important;
        padding: 10px !important;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: #2563EB !important;
        background-color: #F8FAFC !important;
    }
    [data-testid="stFileUploader"] section > button {
        color: #2563EB !important;
        font-weight: 500 !important;
    }

    /* ── Alerts ─────────────────────────────────────────────────────────────── */
    [data-testid="stAlert"] {
        border-radius: 6px !important;
        border: 1px solid #E2E8F0 !important;
        font-size: 13px !important;
        padding: 8px 12px !important;
    }
    .stSuccess, [data-testid="stAlert"][data-baseweb-type="positive"],
    .element-container .stAlert:has(> div[data-testid="stNotificationContentSuccess"]) {
        border-left: 3px solid #10B981 !important;
    }

    /* ── Divider ───────────────────────────────────────────────────────────── */
    hr, [data-testid="stDivider"] {
        border-color: #E2E8F0 !important;
        margin: 12px 0 !important;
    }

    /* ── Global spacing: compact ───────────────────────────────────────────── */
    .element-container {
        margin-bottom: 4px !important;
    }
    [data-testid="stVerticalBlock"] > [data-testid="element-container"] {
        margin-bottom: 4px !important;
    }
    [data-testid="stVerticalBlock"] {
        gap: 4px !important;
    }
    /* Columns gap */
    [data-testid="stHorizontalBlock"] {
        gap: 8px !important;
    }

    /* ── Form (compact) ────────────────────────────────────────────────────── */
    [data-testid="stForm"] {
        background-color: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 10px !important;
        padding: 12px !important;
    }

    /* ── DataFrames / Tables (compact) ─────────────────────────────────────── */
    [data-testid="stDataFrame"] {
        border: 1px solid #E2E8F0 !important;
        border-radius: 6px !important;
        overflow: hidden;
        font-size: 12px !important;
    }
    [data-testid="stDataFrame"] table {
        font-size: 12px !important;
    }
    [data-testid="stDataFrame"] th,
    [data-testid="stDataFrame"] td {
        padding: 4px 8px !important;
        font-size: 12px !important;
        line-height: 1.3 !important;
    }
    /* Glide data-grid cells (Streamlit's internal table renderer) */
    [data-testid="stDataFrame"] .dvn-scroller {
        font-size: 12px !important;
    }

    /* ── Slider (compact) ──────────────────────────────────────────────────── */
    [data-testid="stSlider"] [data-baseweb="slider"] div[role="slider"] {
        background-color: #2563EB !important;
    }
    [data-testid="stSlider"] [data-baseweb="slider"] [data-testid="stTickBar"] {
        background: #2563EB !important;
    }

    /* ── Color picker ──────────────────────────────────────────────────────── */
    [data-testid="stColorPicker"] > div > div {
        border-radius: 6px !important;
    }

    /* ── Scrollbar ──────────────────────────────────────────────────────────── */
    ::-webkit-scrollbar {
        width: 5px;
        height: 5px;
    }
    ::-webkit-scrollbar-track {
        background: #F1F5F9;
    }
    ::-webkit-scrollbar-thumb {
        background: #CBD5E1;
        border-radius: 3px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: #94A3B8;
    }

    /* ── Tooltip ────────────────────────────────────────────────────────────── */
    [data-testid="stTooltipIcon"] {
        color: #94A3B8 !important;
    }

    /* ── Label styling (compact) ───────────────────────────────────────────── */
    .stTextInput label, .stNumberInput label, .stSelectbox label,
    .stMultiSelect label, .stRadio > label, .stCheckbox label,
    [data-testid="stWidgetLabel"] p {
        font-size: 12px !important;
        font-weight: 500 !important;
        color: #334155 !important;
        margin-bottom: 1px !important;
        line-height: 1.2 !important;
    }

    /* ── Reduce widget internal top/bottom padding ─────────────────────────── */
    [data-testid="stWidgetLabel"] {
        margin-bottom: 0 !important;
        padding-bottom: 0 !important;
    }

    /* ── Metric cards ──────────────────────────────────────────────────────── */
    [data-testid="stMetric"] {
        background-color: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 10px !important;
        padding: 10px 12px !important;
    }

    /* ── Plot container ────────────────────────────────────────────────────── */
    [data-testid="stImage"], .stPlotlyChart, [data-testid="stVegaLiteChart"] {
        border-radius: 6px;
        overflow: hidden;
    }

    /* ── Markdown body text ────────────────────────────────────────────────── */
    .stMarkdown p {
        font-size: 13px !important;
        line-height: 1.5 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.title("Sleep EEG Data Visualization")

    uploaded = st.file_uploader("Upload Excel file", type=["xlsx"])
    if not uploaded:
        st.info("Please upload an Excel file to begin.")
        return

    with st.spinner("Loading data..."):
        parsed, mouse_ids = load_and_parse(uploaded)
    st.success(f"Loaded {len(mouse_ids)} mice: {', '.join(mouse_ids[:5])}{'...' if len(mouse_ids) > 5 else ''}")

    with st.sidebar:
        st.header("Group Management")
        if 'groups' not in st.session_state:
            st.session_state.groups = []
        for g in st.session_state.groups:
            if 'gid' not in g:
                g['gid'] = _next_group_id()

        group_file = st.file_uploader("Upload group file (CSV/Excel)", type=["csv", "xlsx", "xls"],
                                      help="Columns: group_name, mice. Supports ranges.",
                                      key="group_file_uploader")
        if group_file is not None:
            file_id = f"{group_file.name}_{group_file.size}"
            if st.session_state.get('_last_group_file_id') != file_id:
                try:
                    gdf = pd.read_csv(group_file) if group_file.name.endswith('.csv') else pd.read_excel(group_file)
                    gdf.columns = [c.strip().lower() for c in gdf.columns]
                    if 'group_name' not in gdf.columns or 'mice' not in gdf.columns:
                        st.error("Need columns: group_name, mice")
                    else:
                        imported = 0
                        for _, row in gdf.iterrows():
                            gname = str(row['group_name']).strip()
                            ids, _ = parse_mouse_ids(str(row['mice']).strip(), mouse_ids)
                            if gname and ids:
                                st.session_state.groups.append(
                                    {'name': gname, 'ids': ids, 'gid': _next_group_id()})
                                imported += 1
                        st.session_state['_last_group_file_id'] = file_id
                        if imported:
                            st.success(f"Imported {imported} group(s)")
                            st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        st.divider()
        with st.form("add_group_form"):
            group_name = st.text_input("Group name", placeholder="e.g., Control")
            mouse_text = st.text_input("Mouse IDs (supports range)", placeholder="e.g., b8281, b8317-b8322")
            group_mice_select = st.multiselect("Or select from list", options=mouse_ids)
            if st.form_submit_button("Add Group") and group_name:
                parsed_ids, parse_errors = parse_mouse_ids(mouse_text, mouse_ids)
                combined = list(dict.fromkeys(parsed_ids + group_mice_select))
                if combined:
                    st.session_state.groups.append(
                        {'name': group_name, 'ids': combined, 'gid': _next_group_id()})
                    st.rerun()
                elif parse_errors:
                    st.error(f"Unrecognized: {', '.join(parse_errors)}")
                else:
                    st.warning("Select at least one mouse.")

        if st.session_state.groups:
            st.subheader("Current Groups")
            action = None
            for i, g in enumerate(st.session_state.groups):
                cols = st.columns([3, 1, 1, 1])
                with cols[0]:
                    st.write(f"**{g['name']}** ({len(g['ids'])} mice)")
                    st.caption(', '.join(g['ids']))
                with cols[1]:
                    if i > 0 and st.button("↑", key=f"up_{i}"):
                        action = ('up', i)
                with cols[2]:
                    if i < len(st.session_state.groups) - 1 and st.button("↓", key=f"dn_{i}"):
                        action = ('down', i)
                with cols[3]:
                    if st.button("✕", key=f"rm_{i}"):
                        action = ('remove', i)
            if action:
                act, idx = action
                if act == 'remove':
                    st.session_state.groups.pop(idx)
                elif act == 'up':
                    lst = st.session_state.groups
                    lst[idx - 1], lst[idx] = lst[idx], lst[idx - 1]
                elif act == 'down':
                    lst = st.session_state.groups
                    lst[idx], lst[idx + 1] = lst[idx + 1], lst[idx]
                st.rerun()

    groups = st.session_state.groups
    if not groups:
        st.warning("Please add at least one group in the sidebar.")
        return

    st.header("Select Plots")
    col1, col2, col3 = st.columns(3)
    with col1:
        plot_bar = st.checkbox("Bar chart: Total sleep time", value=True)
        plot_tc = st.checkbox("Time course: Hourly sleep time", value=True)
    with col2:
        plot_delta = st.checkbox("Delta power density curve", value=True)
        plot_pct = st.checkbox("Delta power percentage curve", value=True)
    with col3:
        plot_spec = st.checkbox("Power spectrum", value=True)

    delta_state = spec_state = 'NREM'
    if plot_delta or plot_pct or plot_spec:
        st.subheader("Plot Options")
        c1, c2 = st.columns(2)
        with c1:
            if plot_delta or plot_pct:
                delta_state = st.selectbox("Delta curves - state", STATES, index=0)
        with c2:
            if plot_spec:
                spec_state = st.selectbox("Power spectrum - state", STATES, index=0)

    if 'show_plots' not in st.session_state:
        st.session_state.show_plots = False
    if st.button("Generate Plots", type="primary"):
        st.session_state.show_plots = True
    if not st.session_state.show_plots:
        return

    st.divider()
    tab_names, tab_keys = [], []
    if plot_bar:
        tab_names.append("Bar Chart")
        tab_keys.append('bar')
    if plot_tc:
        tab_names.append("Time Course")
        tab_keys.append('tc')
    if plot_delta:
        tab_names.append("Delta Density")
        tab_keys.append('delta')
    if plot_pct:
        tab_names.append("Delta %")
        tab_keys.append('pct')
    if plot_spec:
        tab_names.append("Spectrum")
        tab_keys.append('spec')

    if not tab_names:
        st.info("Please select at least one plot type above.")
        return

    plot_tabs = st.tabs(tab_names)
    for tab, tk in zip(plot_tabs, tab_keys):
        with tab:
            render_plot_panel(tk, tk, parsed, groups,
                              delta_state=delta_state, spec_state=spec_state)


if __name__ == '__main__':
    main()
